"""
Libya PMM - Historical Data Excel Export
Exports all historical MEB and commodity data to structured Excel workbook

Filename format: MEB_Historical_Data_YYYYMM-YYYYMM.xlsx
Example: MEB_Historical_Data_202304-202511.xlsx

Generates 12 sheets with dynamic most/least expensive labeling
Auto-deletes old files when new data is added
Includes freeze panes at B2 (row 1 and column A frozen)

Usage:
    python scripts/05_Data_Outputs/export_historical_data.py
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from datetime import datetime
from sqlalchemy import text
from openpyxl.styles import Alignment
from config import PATHS, get_engine

# ============================================================================
# ORDERING CONFIGURATIONS
# ============================================================================

MASTER_ORDER = [
    'AlBayda',
    'Algatroun',
    'AlJufra',
    'AlKhums',
    'AlKufra',
    'Azzawya',
    'Benghazi',
    'Derna',
    'Ejdabia',
    'Ghat',
    'Misrata',
    'Murzuq',
    'Nalut',
    'Sebha',
    'Sirt',
    'Tobruk',
    'Tripoli Center',
    'Ubari',
    'Wadi Alshati',
    'Zliten',
    'Zwara',
    'National Full MEB (Mean)',
    'National Food MEB (Mean)',
    'National Non-Food MEB (Mean)',
    'East',
    'West',
    'South'
]

EAST_ORDER = ['National', 'AlBayda', 'Benghazi', 'Derna', 'Ejdabia', 'AlKufra', 'Tobruk']
WEST_ORDER = ['National', 'AlKhums', 'Azzawya', 'Misrata', 'Nalut', 'Sirt', 'Tripoli Center', 'Zliten', 'Zwara']
SOUTH_ORDER = ['National', 'Algatroun', 'AlJufra', 'Ghat', 'Murzuq', 'Sebha', 'Ubari', 'Wadi Alshati']

FOOD_PRODUCTS_ORDER = [
    'Bread (5pc)', 'Rice (Kg)', 'Pasta (500g)', 'Couscous (Kg)', 
    'Beans (400g)', 'Chicken (Kg)', 'Tuna (200g)', 'Eggs (30pc)',
    'Milk (L)', 'Tomatoes (Kg)', 'Potatoes (Kg)', 'Onions (Kg)',
    'Pepper (Kg)', 'Tomato Paste (400g)', 'Black Tea (250g)',
    'Oil (L)', 'Sugar (Kg)', 'Salt (Kg)'
]

NFI_PRODUCTS_ORDER = [
    'Handwash Soap (Pc)', 'Toothpaste (Pc)', 'Laundry Detergent (L)',
    'Dishwashing Liquid (L)', 'Sanitary Pads (10Pc)', 'Cooking Fuel (11Kg)'
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def add_expense_labels(df, exclude_rows=None):
    """Add (Most/Least Expensive Market) labels based on last month's values"""
    if exclude_rows is None:
        exclude_rows = []
    
    last_col = df.columns[-1]
    muni_mask = ~df.index.isin(exclude_rows)
    muni_values = df.loc[muni_mask, last_col].dropna()
    
    if len(muni_values) == 0:
        return df
    
    highest_muni = muni_values.idxmax()
    lowest_muni = muni_values.idxmin()
    
    new_index = []
    for idx in df.index:
        if idx == highest_muni:
            new_index.append(f"{idx} (Most Expensive Market)")
        elif idx == lowest_muni:
            new_index.append(f"{idx} (Least Expensive Market)")
        else:
            new_index.append(idx)
    
    df.index = new_index
    return df

def reorder_dataframe(df, order_list):
    """Reorder dataframe rows according to order_list"""
    order_dict = {name: i for i, name in enumerate(order_list)}
    missing_items = [idx for idx in df.index if idx not in order_list]
    for i, item in enumerate(missing_items):
        order_dict[item] = len(order_list) + i
    
    df['_sort_order'] = df.index.map(lambda x: order_dict.get(x, 9999))
    df = df.sort_values('_sort_order')
    df = df.drop('_sort_order', axis=1)
    
    return df

# ============================================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================================

def get_master_meb_data(engine, meb_type='full_meb'):
    """Get MEB data for Master sheet with proper ordering"""
    
    query_muni = text(f"""
        SELECT date, municipality, {meb_type}
        FROM municipality_meb
        ORDER BY date, municipality
    """)
    
    query_region = text(f"""
        SELECT date, region, {meb_type}
        FROM regional_meb
        ORDER BY date, region
    """)
    
    query_national = text(f"""
        SELECT date, {meb_type}
        FROM national_meb
        ORDER BY date
    """)
    
    with engine.connect() as conn:
        df_muni = pd.read_sql(query_muni, conn)
        df_muni['date'] = pd.to_datetime(df_muni['date'])
        df_muni[meb_type] = pd.to_numeric(df_muni[meb_type], errors='coerce')
        df_muni.loc[df_muni[meb_type] == 0, meb_type] = None
        pivot_muni = df_muni.pivot(index='municipality', columns='date', values=meb_type)
        
        df_region = pd.read_sql(query_region, conn)
        df_region['date'] = pd.to_datetime(df_region['date'])
        df_region[meb_type] = pd.to_numeric(df_region[meb_type], errors='coerce')
        df_region.loc[df_region[meb_type] == 0, meb_type] = None
        pivot_region = df_region.pivot(index='region', columns='date', values=meb_type)
        
        df_national = pd.read_sql(query_national, conn)
        df_national['date'] = pd.to_datetime(df_national['date'])
        df_national[meb_type] = pd.to_numeric(df_national[meb_type], errors='coerce')
        df_national.loc[df_national[meb_type] == 0, meb_type] = None
        df_national = df_national.set_index('date').T
        
        # FIX: Set label based on MEB type
        if meb_type == 'food_meb':
            df_national.index = ['National Food MEB (Mean)']
        elif meb_type == 'nfi_meb':
            df_national.index = ['National Non-Food MEB (Mean)']
        else:  # full_meb
            df_national.index = ['National Full MEB (Mean)']
    
    combined = pd.concat([df_national, pivot_region, pivot_muni])
    combined = combined.sort_index(axis=1)
    
    # Reorder FIRST (before adding labels)
    combined = reorder_dataframe(combined, MASTER_ORDER)
    
    # THEN add expense labels (so they stay in position)
    exclude_list = [idx for idx in combined.index if 'National' in str(idx) or idx in ['East', 'West', 'South']]
    combined = add_expense_labels(combined, exclude_rows=exclude_list)
    
    combined.columns = [d.strftime('%b-%y') for d in combined.columns]
    combined.index.name = 'Region'
    
    return combined

def get_regional_national_meb(engine, meb_type='full_meb'):
    """Get only regional and national MEB data"""
    
    query_region = text(f"""
        SELECT date, region, {meb_type}
        FROM regional_meb
        ORDER BY date, region
    """)
    
    query_national = text(f"""
        SELECT date, {meb_type}
        FROM national_meb
        ORDER BY date
    """)
    
    with engine.connect() as conn:
        df_region = pd.read_sql(query_region, conn)
        df_region['date'] = pd.to_datetime(df_region['date'])
        df_region[meb_type] = pd.to_numeric(df_region[meb_type], errors='coerce')
        df_region.loc[df_region[meb_type] == 0, meb_type] = None
        pivot_region = df_region.pivot(index='region', columns='date', values=meb_type)
        
        df_national = pd.read_sql(query_national, conn)
        df_national['date'] = pd.to_datetime(df_national['date'])
        df_national[meb_type] = pd.to_numeric(df_national[meb_type], errors='coerce')
        df_national.loc[df_national[meb_type] == 0, meb_type] = None
        df_national = df_national.set_index('date').T
        
        # FIX: Set label based on MEB type
        if meb_type == 'food_meb':
            df_national.index = ['National Food MEB']
        elif meb_type == 'nfi_meb':
            df_national.index = ['National Non-Food MEB']
        else:  # full_meb
            df_national.index = ['National']
    
    combined = pd.concat([df_national, pivot_region])
    combined = combined.sort_index(axis=1)
    combined.columns = [d.strftime('%b-%y') for d in combined.columns]
    combined.index.name = 'Region'
    
    return combined

def get_regional_municipalities_meb(engine, region, order_list, meb_type='food_meb'):
    """Get municipalities from specific region + National + Regional"""
    
    municipalities = [m for m in order_list if m != 'National']
    
    query_muni = text(f"""
        SELECT date, municipality, {meb_type}
        FROM municipality_meb
        WHERE municipality = ANY(:municipalities)
        ORDER BY date, municipality
    """)
    
    query_national = text(f"""
        SELECT date, {meb_type}
        FROM national_meb
        ORDER BY date
    """)
    
    # FIX: Add query for regional data
    query_region = text(f"""
        SELECT date, {meb_type}
        FROM regional_meb
        WHERE region = :region
        ORDER BY date
    """)
    
    with engine.connect() as conn:
        df_muni = pd.read_sql(query_muni, conn, params={'municipalities': municipalities})
        df_muni['date'] = pd.to_datetime(df_muni['date'])
        df_muni[meb_type] = pd.to_numeric(df_muni[meb_type], errors='coerce')
        df_muni.loc[df_muni[meb_type] == 0, meb_type] = None
        pivot_muni = df_muni.pivot(index='municipality', columns='date', values=meb_type)
        
        df_national = pd.read_sql(query_national, conn)
        df_national['date'] = pd.to_datetime(df_national['date'])
        df_national[meb_type] = pd.to_numeric(df_national[meb_type], errors='coerce')
        df_national.loc[df_national[meb_type] == 0, meb_type] = None
        df_national = df_national.set_index('date').T
        
        # FIX: Set label based on MEB type
        if meb_type == 'food_meb':
            df_national.index = ['National Food MEB']
        elif meb_type == 'nfi_meb':
            df_national.index = ['National Non-Food MEB']
        else:  # full_meb
            df_national.index = ['National Full MEB']
        
        # FIX: Get regional data
        df_region = pd.read_sql(query_region, conn, params={'region': region})
        df_region['date'] = pd.to_datetime(df_region['date'])
        df_region[meb_type] = pd.to_numeric(df_region[meb_type], errors='coerce')
        df_region.loc[df_region[meb_type] == 0, meb_type] = None
        df_region = df_region.set_index('date').T
        
        # FIX: Set label based on MEB type
        if meb_type == 'food_meb':
            df_region.index = [f'{region} Food MEB']
        elif meb_type == 'nfi_meb':
            df_region.index = [f'{region} Non-Food MEB']
        else:  # full_meb
            df_region.index = [f'{region} Full MEB']
    
    # FIX: Combine municipalities + National + Regional (at end)
    combined = pd.concat([pivot_muni, df_national, df_region])
    combined = combined.sort_index(axis=1)
    
    # Reorder FIRST (before adding labels)
    combined = reorder_dataframe(combined, order_list)
    
    # THEN add expense labels (so they stay in position) - exclude National and Regional
    exclude_list = [idx for idx in combined.index if 'National' in str(idx) or any(r in str(idx) for r in ['East', 'West', 'South']) and 'MEB' in str(idx)]
    combined = add_expense_labels(combined, exclude_rows=exclude_list)
    
    combined.columns = [d.strftime('%b-%y') for d in combined.columns]
    combined.index.name = 'Region'
    
    return combined

def get_commodity_data(engine):
    """Get commodity data with proper ordering"""
    
    query = text("""
        SELECT 
            date,
            product_name,
            category,
            admin_name,
            average_price
        FROM products
        ORDER BY admin_name, category, product_name, date
    """)
    
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
        df['date'] = pd.to_datetime(df['date'])
        df['average_price'] = df['average_price'].astype(float)
    
    result_rows = []
    
    for admin in ['Libya', 'East', 'West', 'South']:
        admin_data = df[df['admin_name'] == admin]
        
        for products_list, category in [(FOOD_PRODUCTS_ORDER, 'Food'), 
                                         (NFI_PRODUCTS_ORDER, 'Non-Food')]:
            for product_name in products_list:
                product_data = admin_data[admin_data['product_name'] == product_name]
                
                if not product_data.empty:
                    row_data = {
                        'Region': 'National' if admin == 'Libya' else admin,
                        'Type': category,
                        'Item': product_name
                    }
                    
                    for _, row in product_data.iterrows():
                        date_str = row['date'].strftime('%b-%y')
                        row_data[date_str] = row['average_price']
                    
                    result_rows.append(row_data)
    
    df_result = pd.DataFrame(result_rows)
    
    fixed_cols = ['Region', 'Type', 'Item']
    date_cols = sorted([c for c in df_result.columns if c not in fixed_cols], 
                       key=lambda x: datetime.strptime(x, '%b-%y'))
    df_result = df_result[fixed_cols + date_cols]
    
    return df_result

# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def export_historical_data():
    """Export all historical data to Excel workbook with dynamic date range in filename"""
    
    engine = get_engine()
    
    print("="*70)
    print("LIBYA PMM - HISTORICAL DATA EXPORT")
    print("="*70)
    
    # Get date range from data (use Full MEB to determine range)
    print("\nDetermining date range from data...")
    temp_df = get_master_meb_data(engine, 'full_meb')
    
    # Get start and end dates from column names (format: 'Mon-YY')
    date_columns = [col for col in temp_df.columns if col != 'Region']
    start_date_str = date_columns[0]  # First column (e.g., 'Apr-23')
    end_date_str = date_columns[-1]   # Last column (e.g., 'Nov-25')
    
    # Parse to get YYYYMM format
    start_date = pd.to_datetime(start_date_str, format='%b-%y')
    end_date = pd.to_datetime(end_date_str, format='%b-%y')
    
    start_yyyymm = start_date.strftime('%Y%m')
    end_yyyymm = end_date.strftime('%Y%m')
    
    date_range_str = f"{start_yyyymm}-{end_yyyymm}"
    
    print(f"   Date range: {start_date.strftime('%B %Y')} to {end_date.strftime('%B %Y')}")
    print(f"   Filename tag: {date_range_str}")
    
    # Output file with date range
    output_file = PATHS['master_data'] / f'MEB_Historical_Data_{date_range_str}.xlsx'
    
    # Clean up old files with different date ranges
    print(f"\n🗑️  Cleaning up old files...")
    old_files = list(PATHS['master_data'].glob('MEB_Historical_Data_*.xlsx'))
    
    deleted_count = 0
    for old_file in old_files:
        if old_file != output_file:
            print(f"   Deleting old file: {old_file.name}")
            old_file.unlink()
            deleted_count += 1
    
    if deleted_count > 0:
        print(f"   ✓ Deleted {deleted_count} old file(s)")
    else:
        print(f"   No old files to delete")
    
    print(f"\nCreating workbook: {output_file.name}")
    print(f"Location: {PATHS['master_data']}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        print(f"\n1. Generating: Full MEB - Master...")
        df1 = get_master_meb_data(engine, 'full_meb')
        df1.to_excel(writer, sheet_name='Full MEB - Master')
        
        print(f"2. Generating: Food MEB - Master...")
        df2 = get_master_meb_data(engine, 'food_meb')
        df2.to_excel(writer, sheet_name='Food MEB - Master')
        
        print(f"3. Generating: NFI MEB - Master...")
        df3 = get_master_meb_data(engine, 'nfi_meb')
        df3.to_excel(writer, sheet_name='NFI MEB - Master')
        
        print(f"4. Generating: Commodities - Master...")
        df4 = get_commodity_data(engine)
        df4.to_excel(writer, sheet_name='Commodities - Master', index=False)
        
        print(f"5. Generating: Food MEB - National & Regions...")
        df5 = get_regional_national_meb(engine, 'food_meb')
        df5.to_excel(writer, sheet_name='Food MEB - Nat & Reg')
        
        print(f"6. Generating: NFI MEB - National & Regions...")
        df6 = get_regional_national_meb(engine, 'nfi_meb')
        df6.to_excel(writer, sheet_name='NFI MEB - Nat & Reg')
        
        print(f"7. Generating: Food MEB - East...")
        df7 = get_regional_municipalities_meb(engine, 'East', EAST_ORDER, 'food_meb')
        df7.to_excel(writer, sheet_name='Food MEB - East')
        
        print(f"8. Generating: NFI MEB - East...")
        df8 = get_regional_municipalities_meb(engine, 'East', EAST_ORDER, 'nfi_meb')
        df8.to_excel(writer, sheet_name='NFI MEB - East')
        
        print(f"9. Generating: Food MEB - West...")
        df9 = get_regional_municipalities_meb(engine, 'West', WEST_ORDER, 'food_meb')
        df9.to_excel(writer, sheet_name='Food MEB - West')
        
        print(f"10. Generating: NFI MEB - West...")
        df10 = get_regional_municipalities_meb(engine, 'West', WEST_ORDER, 'nfi_meb')
        df10.to_excel(writer, sheet_name='NFI MEB - West')
        
        print(f"11. Generating: Food MEB - South...")
        df11 = get_regional_municipalities_meb(engine, 'South', SOUTH_ORDER, 'food_meb')
        df11.to_excel(writer, sheet_name='Food MEB - South')
        
        print(f"12. Generating: NFI MEB - South...")
        df12 = get_regional_municipalities_meb(engine, 'South', SOUTH_ORDER, 'nfi_meb')
        df12.to_excel(writer, sheet_name='NFI MEB - South')
    
    # Apply freeze panes, borders, and adjust column widths
    print(f"\nApplying borders, freeze panes, and adjusting column widths...")
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    
    wb = load_workbook(output_file)
    
    # Define border styles (matching master_data.py)
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    thick_top_border = Border(
        top=Side(style='thick', color='000000'),
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    thick_bottom_border = Border(
        bottom=Side(style='thick', color='000000'),
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9')
    )
    
    # Special borders for Nat & Reg sheets (only top and bottom thick borders)
    thick_top_only = Border(top=Side(style='thick', color='000000'))
    thick_bottom_only = Border(bottom=Side(style='thick', color='000000'))
    
    # Regular font (not bold)
    regular_font = Font(bold=False)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Special freeze for Commodities sheet (3 columns: Region, Type, Item)
        if sheet_name == 'Commodities - Master':
            ws.freeze_panes = 'D2'  # Freeze first 3 columns (A, B, C) and row 1
            print(f"   ✓ {sheet_name}: Freeze at D2 (3 columns frozen)")
        else:
            ws.freeze_panes = 'B2'  # Standard freeze (1 column and row 1)
            print(f"   ✓ {sheet_name}: Freeze at B2")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set adjusted width (add padding)
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Skip border formatting for Commodities
        if sheet_name == 'Commodities - Master':
            continue
        
        # Special formatting for "Nat & Reg" sheets
        if 'Nat & Reg' in sheet_name:
            ws.column_dimensions['A'].width = 35
            
            # Remove bold and apply only thick top and bottom borders
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font  # Remove bold
                    
                    # Thick border on top of row 2
                    if row_idx == 2:
                        cell.border = thick_top_only
                    # Thick border on bottom of row 5 (last data row)
                    elif row_idx == 5:
                        cell.border = thick_bottom_only
                    else:
                        cell.border = None  # No borders for rows 3 and 4
                    
                    if col_idx == 1:  # Left align column A
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            print(f"   ✓ {sheet_name}: Thick borders on row 2 (top) and row 5 (bottom) only")
            continue
        
        # Apply borders to Master and Regional MEB sheets
        ws.column_dimensions['A'].width = 35  # Set column A width for long names
        
        # Find summary section rows
        national_row = None
        south_row = None
        last_region_row = None
        
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and 'National' in str(cell_value):
                if national_row is None:
                    national_row = row_idx
            
            if cell_value and cell_value == 'South':
                south_row = row_idx
            
            if cell_value and ((' Full MEB' in str(cell_value) or ' Food MEB' in str(cell_value) or ' Non-Food MEB' in str(cell_value))
                              and any(region in str(cell_value) for region in ['East', 'West', 'South'])):
                last_region_row = row_idx
        
        # Apply borders, white fill, and remove bold from all data cells
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = white_fill
                cell.border = thin_border
                cell.font = regular_font  # Remove bold
                if col_idx == 1:  # Left align column A
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Apply thick borders for summary sections
        if national_row:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=national_row, column=col_idx).border = thick_top_border
        
        if south_row:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=south_row, column=col_idx).border = thick_bottom_border
        
        if last_region_row and not south_row:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=last_region_row, column=col_idx).border = thick_bottom_border
    
    wb.save(output_file)
    print(f"   ✓ Borders applied to Master and Regional sheets")
    print(f"   ✓ Column widths adjusted for all sheets")
    print(f"   ✓ Column A left-aligned and not bold")
    
    print(f"\n" + "="*70)
    print("EXPORT COMPLETE")
    print("="*70)
    print(f"\n✓ Workbook saved: {output_file.name}")
    print(f"  12 sheets generated")
    print(f"  File size: {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  Location: {output_file}")
    
    return output_file

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Export all historical MEB and commodity data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export all historical data
  python scripts/05_Data_Outputs/export_historical_data.py

Output:
  MEB_Historical_Data_YYYYMM-YYYYMM.xlsx with 12 sheets:
  Example: MEB_Historical_Data_202304-202511.xlsx
  
  Sheets:
  - Full MEB - Master
  - Food MEB - Master
  - NFI MEB - Master
  - Commodities - Master
  - Food MEB - Nat & Reg
  - NFI MEB - Nat & Reg
  - Food MEB - East
  - NFI MEB - East
  - Food MEB - West
  - NFI MEB - West
  - Food MEB - South
  - NFI MEB - South

Features:
  - Dynamic date range in filename (auto-updates)
  - Auto-deletes old files
  - Freeze panes: B2 for most sheets, D2 for Commodities (3 columns frozen)
  - Auto-adjusted column widths for better readability

Prerequisites:
  Database must be loaded with PMM data
        """
    )
    
    args = parser.parse_args()
    
    try:
        output_file = export_historical_data()
        print(f"\n✅ Ready to use in Excel or PowerPoint!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)