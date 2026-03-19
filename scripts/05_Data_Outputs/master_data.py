"""
Libya PMM - Historical MEB Export (Master + Regional Sheets with MoM% and YoY%)
Exports to TWO separate files with dynamic date ranges:

1. MoM_MEB_TrendAnalysis_YYYYMM-YYYYMM.xlsx (12 sheets):
   Example: MoM_MEB_TrendAnalysis_202304-202511.xlsx
   - Master MoM% (Full, Food, NFI) - All municipalities + regions
   - East MoM% (Full, Food, NFI) - East municipalities only
   - West MoM% (Full, Food, NFI) - West municipalities only
   - South MoM% (Full, Food, NFI) - South municipalities only

2. YoY_MEB_TrendAnalysis_YYYYMM-YYYYMM.xlsx (12 sheets):
   Example: YoY_MEB_TrendAnalysis_202304-202511.xlsx
   - Master YoY% (Full, Food, NFI) - All municipalities + regions
   - East YoY% (Full, Food, NFI) - East municipalities only
   - West YoY% (Full, Food, NFI) - West municipalities only
   - South YoY% (Full, Food, NFI) - South municipalities only

Filename includes date range: YYYYMM-YYYYMM (start month - end month)
Updates automatically when new data is added to database.

Format:
- Row 1: MEB value (digits only, no "LYD")
- Row 2: MoM% or YoY% with ▲/▼ indicators
- Colors: Red for increase, Green for decrease

Usage:
    python scripts/05_Data_Outputs/master_data.py
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from datetime import datetime
from sqlalchemy import text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import PATHS, get_engine

# ============================================================================
# MASTER ORDER
# ============================================================================

MASTER_ORDER = [
    'AlBayda', 'Algatroun', 'AlJufra', 'AlKhums', 'AlKufra', 'Azzawya',
    'Benghazi', 'Derna', 'Ejdabia', 'Ghat', 'Misrata', 'Murzuq', 'Nalut',
    'Sebha', 'Sirt', 'Tobruk', 'Tripoli Center', 'Ubari', 'Wadi Alshati',
    'Zliten', 'Zwara',
    'National Full MEB (Mean)',
    'East', 'West', 'South'
]

# Regional municipality lists
EAST_MUNICIPALITIES = ['AlBayda', 'Benghazi', 'Derna', 'Ejdabia', 'AlKufra', 'Tobruk']
WEST_MUNICIPALITIES = ['AlKhums', 'Azzawya', 'Misrata', 'Nalut', 'Sirt', 'Tripoli Center', 'Zliten', 'Zwara']
SOUTH_MUNICIPALITIES = ['Algatroun', 'AlJufra', 'Ghat', 'Murzuq', 'Sebha', 'Ubari', 'Wadi Alshati']

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def add_expense_labels(df, exclude_rows=None):
    """Add (Most/Least Expensive Market) labels"""
    if exclude_rows is None:
        exclude_rows = []
    
    last_col = df.columns[-1]
    muni_mask = ~df.index.isin(exclude_rows)
    muni_values = df.loc[muni_mask, last_col].dropna()
    
    if len(muni_values) == 0:
        return df
    
    highest_muni = muni_values.idxmax()
    lowest_muni = muni_values.idxmin()
    
    new_index = []
    for idx in df.index:
        if idx == highest_muni:
            new_index.append(f"{idx} (Most Expensive)")
        elif idx == lowest_muni:
            new_index.append(f"{idx} (Least Expensive)")
        else:
            new_index.append(idx)
    
    df.index = new_index
    return df

def reorder_dataframe(df, order_list):
    """Reorder dataframe rows"""
    order_dict = {name: i for i, name in enumerate(order_list)}
    missing_items = [idx for idx in df.index if idx not in order_list]
    for i, item in enumerate(missing_items):
        order_dict[item] = len(order_list) + i
    
    df['_sort_order'] = df.index.map(lambda x: order_dict.get(x, 9999))
    df = df.sort_values('_sort_order')
    df = df.drop('_sort_order', axis=1)
    
    return df

def calculate_mom_pct(df):
    """Calculate Month-over-Month % change"""
    mom_df = df.copy()
    
    for idx in mom_df.index:
        row_values = mom_df.loc[idx]
        mom_values = []
        
        for i in range(len(row_values)):
            if i == 0:
                mom_values.append(None)
            else:
                current = row_values.iloc[i]
                previous = row_values.iloc[i-1]
                
                if pd.notna(current) and pd.notna(previous) and previous > 0:
                    mom_pct = ((current - previous) / previous) * 100
                    mom_values.append(mom_pct)
                else:
                    mom_values.append(None)
        
        mom_df.loc[idx] = mom_values
    
    return mom_df

def calculate_yoy_pct(df):
    """Calculate Year-over-Year % change (12 months)"""
    yoy_df = df.copy()
    
    for idx in yoy_df.index:
        row_values = yoy_df.loc[idx]
        yoy_values = []
        
        for i in range(len(row_values)):
            if i < 12:
                yoy_values.append(None)  # Need 12 months history
            else:
                current = row_values.iloc[i]
                year_ago = row_values.iloc[i-12]
                
                if pd.notna(current) and pd.notna(year_ago) and year_ago > 0:
                    yoy_pct = ((current - year_ago) / year_ago) * 100
                    yoy_values.append(yoy_pct)
                else:
                    yoy_values.append(None)
        
        yoy_df.loc[idx] = yoy_values
    
    return yoy_df

# ============================================================================
# DATA EXTRACTION
# ============================================================================

def get_master_meb_data(engine, meb_type='full_meb'):
    """Get MEB data for Master sheet (all municipalities + regions + national)"""
    
    query_muni = text(f"""
        SELECT date, municipality, {meb_type}
        FROM municipality_meb
        ORDER BY date, municipality
    """)
    
    query_region = text(f"""
        SELECT date, region, {meb_type}
        FROM regional_meb
        ORDER BY date, region
    """)
    
    query_national = text(f"""
        SELECT date, {meb_type}
        FROM national_meb
        ORDER BY date
    """)
    
    with engine.connect() as conn:
        df_muni = pd.read_sql(query_muni, conn)
        df_muni['date'] = pd.to_datetime(df_muni['date'])
        df_muni[meb_type] = pd.to_numeric(df_muni[meb_type], errors='coerce')
        df_muni.loc[df_muni[meb_type] == 0, meb_type] = None
        pivot_muni = df_muni.pivot(index='municipality', columns='date', values=meb_type)
        
        df_region = pd.read_sql(query_region, conn)
        df_region['date'] = pd.to_datetime(df_region['date'])
        df_region[meb_type] = pd.to_numeric(df_region[meb_type], errors='coerce')
        df_region.loc[df_region[meb_type] == 0, meb_type] = None
        pivot_region = df_region.pivot(index='region', columns='date', values=meb_type)
        
        df_national = pd.read_sql(query_national, conn)
        df_national['date'] = pd.to_datetime(df_national['date'])
        df_national[meb_type] = pd.to_numeric(df_national[meb_type], errors='coerce')
        df_national.loc[df_national[meb_type] == 0, meb_type] = None
        df_national = df_national.set_index('date').T
        
        # Set appropriate label based on MEB type
        if meb_type == 'full_meb':
            df_national.index = ['National Full MEB (Mean)']
        elif meb_type == 'food_meb':
            df_national.index = ['National Food MEB']
        elif meb_type == 'nfi_meb':
            df_national.index = ['National Non-Food MEB']
    
    combined = pd.concat([df_national, pivot_region, pivot_muni])
    combined = combined.sort_index(axis=1)
    
    # Create dynamic order list based on MEB type
    municipality_list = [
        'AlBayda', 'Algatroun', 'AlJufra', 'AlKhums', 'AlKufra', 'Azzawya',
        'Benghazi', 'Derna', 'Ejdabia', 'Ghat', 'Misrata', 'Murzuq', 'Nalut',
        'Sebha', 'Sirt', 'Tobruk', 'Tripoli Center', 'Ubari', 'Wadi Alshati',
        'Zliten', 'Zwara'
    ]
    
    # Set appropriate national label and create order
    if meb_type == 'full_meb':
        national_label = 'National Full MEB (Mean)'
    elif meb_type == 'food_meb':
        national_label = 'National Food MEB'
    elif meb_type == 'nfi_meb':
        national_label = 'National Non-Food MEB'
    
    # Order: municipalities first, then National, then regions
    dynamic_order = municipality_list + [national_label, 'East', 'West', 'South']
    
    combined = reorder_dataframe(combined, dynamic_order)
    
    combined = add_expense_labels(combined, exclude_rows=[national_label, 'East', 'West', 'South'])
    
    # Calculate both MoM and YoY
    mom_df = calculate_mom_pct(combined)
    yoy_df = calculate_yoy_pct(combined)
    
    # Format column names
    combined.columns = [d.strftime('%b-%y') for d in combined.columns]
    mom_df.columns = [d.strftime('%b-%y') for d in mom_df.columns]
    yoy_df.columns = [d.strftime('%b-%y') for d in yoy_df.columns]
    
    return combined, mom_df, yoy_df

def get_regional_meb_data(engine, meb_type='full_meb', region='East'):
    """Get MEB data for a specific region (municipalities only)"""
    
    # Select municipalities based on region
    if region == 'East':
        municipalities = EAST_MUNICIPALITIES
    elif region == 'West':
        municipalities = WEST_MUNICIPALITIES
    elif region == 'South':
        municipalities = SOUTH_MUNICIPALITIES
    else:
        raise ValueError(f"Unknown region: {region}")
    
    query_muni = text(f"""
        SELECT date, municipality, {meb_type}
        FROM municipality_meb
        WHERE municipality = ANY(:municipalities)
        ORDER BY date, municipality
    """)
    
    # Query for national MEB
    query_national = text(f"""
        SELECT date, {meb_type}
        FROM national_meb
        ORDER BY date
    """)
    
    # Query for regional MEB
    query_region = text(f"""
        SELECT date, {meb_type}
        FROM regional_meb
        WHERE region = :region
        ORDER BY date
    """)
    
    with engine.connect() as conn:
        df_muni = pd.read_sql(query_muni, conn, params={'municipalities': municipalities})
        df_muni['date'] = pd.to_datetime(df_muni['date'])
        df_muni[meb_type] = pd.to_numeric(df_muni[meb_type], errors='coerce')
        df_muni.loc[df_muni[meb_type] == 0, meb_type] = None
        pivot_muni = df_muni.pivot(index='municipality', columns='date', values=meb_type)
        
        # Get national data
        df_national = pd.read_sql(query_national, conn)
        df_national['date'] = pd.to_datetime(df_national['date'])
        df_national[meb_type] = pd.to_numeric(df_national[meb_type], errors='coerce')
        df_national.loc[df_national[meb_type] == 0, meb_type] = None
        df_national = df_national.set_index('date').T
        
        # Set appropriate national label based on MEB type
        if meb_type == 'full_meb':
            df_national.index = ['National Full MEB']
        elif meb_type == 'food_meb':
            df_national.index = ['National Food MEB']
        elif meb_type == 'nfi_meb':
            df_national.index = ['National Non-Food MEB']
        
        # Get regional data
        df_region = pd.read_sql(query_region, conn, params={'region': region})
        df_region['date'] = pd.to_datetime(df_region['date'])
        df_region[meb_type] = pd.to_numeric(df_region[meb_type], errors='coerce')
        df_region.loc[df_region[meb_type] == 0, meb_type] = None
        df_region = df_region.set_index('date').T
        
        # Set appropriate regional label based on MEB type
        if meb_type == 'full_meb':
            df_region.index = [f'{region} Full MEB']
        elif meb_type == 'food_meb':
            df_region.index = [f'{region} Food MEB']
        elif meb_type == 'nfi_meb':
            df_region.index = [f'{region} Non-Food MEB']
    
    # Reorder municipalities
    pivot_muni = pivot_muni.reindex(municipalities)
    pivot_muni = pivot_muni.sort_index(axis=1)
    pivot_muni = add_expense_labels(pivot_muni, exclude_rows=[])
    
    # Combine: Municipalities + National + Regional
    combined = pd.concat([pivot_muni, df_national, df_region])
    combined = combined.sort_index(axis=1)
    
    # Calculate both MoM and YoY from combined
    mom_df = calculate_mom_pct(combined)
    yoy_df = calculate_yoy_pct(combined)
    
    # Format column names
    combined.columns = [d.strftime('%b-%y') for d in combined.columns]
    mom_df.columns = [d.strftime('%b-%y') for d in mom_df.columns]
    yoy_df.columns = [d.strftime('%b-%y') for d in yoy_df.columns]
    
    return combined, mom_df, yoy_df

# ============================================================================
# CREATE COMBINED DATAFRAME
# ============================================================================

def create_combined_dataframe(values_df, change_df, change_type='MoM'):
    """Create dataframe with alternating value and change% rows"""
    result_rows = []
    
    for idx in values_df.index:
        # Value row - just the number
        value_row = [idx]
        for col in values_df.columns:
            val = values_df.loc[idx, col]
            if pd.notna(val):
                value_row.append(f"{val:.2f}")  # No "LYD" prefix
            else:
                value_row.append("")
        result_rows.append(value_row)
        
        # Change% row - empty region name (will be merged with above in Excel)
        change_row = [""]  # Empty string
        for col in change_df.columns:
            change_val = change_df.loc[idx, col]
            if pd.notna(change_val):
                if change_val > 0:
                    change_row.append(f"▲ +{change_val:.1f}%")
                elif change_val < 0:
                    change_row.append(f"▼ {change_val:.1f}%")
                else:
                    change_row.append(f"{change_val:.1f}%")
            else:
                change_row.append("")
        result_rows.append(change_row)
    
    columns = ['Region'] + list(values_df.columns)
    return pd.DataFrame(result_rows, columns=columns)

# ============================================================================
# APPLY FORMATTING
# ============================================================================

def apply_formatting(workbook_path):
    """Apply color formatting, merge region cells, set backgrounds, column widths, and borders"""
    
    wb = load_workbook(workbook_path)
    
    # Red and green fonts
    red_font = Font(name='Aptos Narrow', size=11, color='C00000')
    green_font = Font(name='Aptos Narrow', size=11, color='00B050')
    
    # White fill for data cells
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Thick black borders for national/regional summary section
    thick_top_border = Border(
        top=Side(style='thick', color='000000'),
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    thick_bottom_border = Border(
        bottom=Side(style='thick', color='000000'),
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Set column width for Region column (Column A) to fit long names
        ws.column_dimensions['A'].width = 35
        
        # Find where summary section starts (National row) and ends (South row)
        # In Master sheets: National, East, West, South (need borders before National and after South)
        # In Regional sheets: National, Region (need borders before National and after Region)
        national_row = None
        south_row = None
        last_region_row = None
        
        for row_idx in range(2, ws.max_row + 1, 2):  # Check every value row
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and 'National' in str(cell_value):
                if national_row is None:  # First occurrence
                    national_row = row_idx
            
            # Check for South in master sheets
            if cell_value and cell_value == 'South':
                south_row = row_idx
            
            # Check for regional summary rows (East/West/South Full/Food/Non-Food MEB at the end)
            if cell_value and ((' Full MEB' in str(cell_value) or ' Food MEB' in str(cell_value) or ' Non-Food MEB' in str(cell_value)) 
                              and any(region in str(cell_value) for region in ['East', 'West', 'South'])):
                last_region_row = row_idx
        
        # Merge region cells (Column A) for each municipality
        row_idx = 2
        while row_idx <= ws.max_row:
            # Merge cells A(row_idx) and A(row_idx+1)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx+1, end_column=1)
            
            # Center align the merged cell vertically, left align horizontally
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
            
            # Apply white background to both rows
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = white_fill
                ws.cell(row=row_idx + 1, column=col_idx).fill = white_fill
                
                # Apply thin borders to all data cells
                ws.cell(row=row_idx, column=col_idx).border = thin_border
                ws.cell(row=row_idx + 1, column=col_idx).border = thin_border
            
            # Apply special borders for summary section
            if national_row and row_idx == national_row:
                # Thick black border on top of National row (first summary row)
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).border = thick_top_border
            
            # For master sheets: thick border after South
            if south_row and row_idx == south_row:
                # Thick border on bottom of South row (last region in master sheets)
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx + 1, column=col_idx).border = thick_bottom_border
            
            # For regional sheets: thick border after the regional summary row
            if last_region_row and row_idx == last_region_row and not south_row:
                # Only apply if there's no South row (meaning this is a regional sheet, not master)
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx + 1, column=col_idx).border = thick_bottom_border
            
            row_idx += 2  # Move to next municipality (skip the change% row)
        
        # Apply color formatting to change% rows
        # Every odd row after header (3, 5, 7, ...) is a change% row
        for row_idx in range(3, ws.max_row + 1, 2):  # Start at 3 (first change% row), step by 2
            # Check column B onwards (skip Region column)
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('▲'):  # Increase
                        cell.font = red_font
                    elif cell.value.startswith('▼'):  # Decrease
                        cell.font = green_font
        
        # Freeze panes at B2 (keeps row 1 and column A frozen)
        ws.freeze_panes = 'B2'
    
    wb.save(workbook_path)

# ============================================================================
# MAIN EXPORT
# ============================================================================

# This file contains just the replacement for the export function
# Will be inserted into the main file

def export_historical_meb():
    """Export MoM and YoY data to separate Excel files with date ranges in filenames"""
    
    engine = get_engine()
    
    print("="*70)
    print("LIBYA PMM - HISTORICAL MEB EXPORT (MoM% & YoY%)")
    print("="*70)
    
    # Create output directory if needed
    PATHS['master_data'].mkdir(parents=True, exist_ok=True)
    
    # Get date range from data (use Full MEB to determine range)
    print("\nDetermining date range from data...")
    val, mom, yoy = get_master_meb_data(engine, 'full_meb')
    
    # Get start and end dates from column names (format: 'Mon-YY')
    date_columns = val.columns.tolist()
    start_date_str = date_columns[0]  # First column (e.g., 'Apr-23')
    end_date_str = date_columns[-1]   # Last column (e.g., 'Nov-25')
    
    # Parse to get YYYYMM format
    start_date = pd.to_datetime(start_date_str, format='%b-%y')
    end_date = pd.to_datetime(end_date_str, format='%b-%y')
    
    start_yyyymm = start_date.strftime('%Y%m')
    end_yyyymm = end_date.strftime('%Y%m')
    
    date_range_str = f"{start_yyyymm}-{end_yyyymm}"
    
    print(f"   Date range: {start_date.strftime('%B %Y')} to {end_date.strftime('%B %Y')}")
    print(f"   Filename tag: {date_range_str}")
    
    # Two separate output files with date ranges
    mom_output = PATHS['master_data'] / f'MoM_MEB_TrendAnalysis_{date_range_str}.xlsx'
    yoy_output = PATHS['master_data'] / f'YoY_MEB_TrendAnalysis_{date_range_str}.xlsx'
    
    # Clean up old files with different date ranges
    print(f"\n🗑️  Cleaning up old files...")
    old_mom_files = list(PATHS['master_data'].glob('MoM_MEB_TrendAnalysis_*.xlsx'))
    old_yoy_files = list(PATHS['master_data'].glob('YoY_MEB_TrendAnalysis_*.xlsx'))
    
    deleted_count = 0
    for old_file in old_mom_files + old_yoy_files:
        if old_file != mom_output and old_file != yoy_output:
            print(f"   Deleting old file: {old_file.name}")
            old_file.unlink()
            deleted_count += 1
    
    if deleted_count > 0:
        print(f"   ✓ Deleted {deleted_count} old file(s)")
    else:
        print(f"   No old files to delete")
    
    # ========== MoM FILE (12 sheets) ==========
    print(f"\n{'='*70}")
    print("CREATING MoM FILE")
    print(f"{'='*70}")
    print(f"\nCreating: {mom_output.name}")
    print(f"Location: {PATHS['master_data']}")
    print("\nGenerating MoM sheets...")
    
    with pd.ExcelWriter(mom_output, engine='openpyxl') as writer:
        
        # MASTER MoM (already fetched above for date range)
        print(f"\n  MASTER MoM SHEETS:")
        print(f"  1. Master MoM - Full MEB...")
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='Master MoM - Full', index=False)
        
        print(f"  2. Master MoM - Food MEB...")
        val, mom, yoy = get_master_meb_data(engine, 'food_meb')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='Master MoM - Food', index=False)
        
        print(f"  3. Master MoM - NFI MEB...")
        val, mom, yoy = get_master_meb_data(engine, 'nfi_meb')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='Master MoM - NFI', index=False)
        
        # EAST MoM
        print(f"\n  EAST MoM SHEETS:")
        print(f"  4. East MoM - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'East')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='East MoM - Full', index=False)
        
        print(f"  5. East MoM - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'East')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='East MoM - Food', index=False)
        
        print(f"  6. East MoM - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'East')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='East MoM - NFI', index=False)
        
        # WEST MoM
        print(f"\n  WEST MoM SHEETS:")
        print(f"  7. West MoM - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'West')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='West MoM - Full', index=False)
        
        print(f"  8. West MoM - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'West')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='West MoM - Food', index=False)
        
        print(f"  9. West MoM - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'West')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='West MoM - NFI', index=False)
        
        # SOUTH MoM
        print(f"\n  SOUTH MoM SHEETS:")
        print(f"  10. South MoM - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'South')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='South MoM - Full', index=False)
        
        print(f"  11. South MoM - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'South')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='South MoM - Food', index=False)
        
        print(f"  12. South MoM - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'South')
        create_combined_dataframe(val, mom, 'MoM').to_excel(writer, sheet_name='South MoM - NFI', index=False)
    
    # Apply formatting to MoM file
    print(f"\nApplying color formatting to MoM file...")
    apply_formatting(mom_output)
    
    print(f"\n✓ MoM file saved: {mom_output.name}")
    print(f"  Size: {mom_output.stat().st_size / 1024:.1f} KB")
    
    # ========== YoY FILE (12 sheets) ==========
    print(f"\n{'='*70}")
    print("CREATING YoY FILE")
    print(f"{'='*70}")
    print(f"\nCreating: {yoy_output.name}")
    print(f"Location: {PATHS['master_data']}")
    print("\nGenerating YoY sheets...")
    
    with pd.ExcelWriter(yoy_output, engine='openpyxl') as writer:
        
        # MASTER YoY (use data already fetched for date range)
        print(f"\n  MASTER YoY SHEETS:")
        print(f"  1. Master YoY - Full MEB...")
        val, mom, yoy = get_master_meb_data(engine, 'full_meb')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='Master YoY - Full', index=False)
        
        print(f"  2. Master YoY - Food MEB...")
        val, mom, yoy = get_master_meb_data(engine, 'food_meb')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='Master YoY - Food', index=False)
        
        print(f"  3. Master YoY - NFI MEB...")
        val, mom, yoy = get_master_meb_data(engine, 'nfi_meb')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='Master YoY - NFI', index=False)
        
        # EAST YoY
        print(f"\n  EAST YoY SHEETS:")
        print(f"  4. East YoY - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'East')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='East YoY - Full', index=False)
        
        print(f"  5. East YoY - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'East')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='East YoY - Food', index=False)
        
        print(f"  6. East YoY - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'East')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='East YoY - NFI', index=False)
        
        # WEST YoY
        print(f"\n  WEST YoY SHEETS:")
        print(f"  7. West YoY - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'West')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='West YoY - Full', index=False)
        
        print(f"  8. West YoY - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'West')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='West YoY - Food', index=False)
        
        print(f"  9. West YoY - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'West')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='West YoY - NFI', index=False)
        
        # SOUTH YoY
        print(f"\n  SOUTH YoY SHEETS:")
        print(f"  10. South YoY - Full MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'full_meb', 'South')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='South YoY - Full', index=False)
        
        print(f"  11. South YoY - Food MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'food_meb', 'South')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='South YoY - Food', index=False)
        
        print(f"  12. South YoY - NFI MEB...")
        val, mom, yoy = get_regional_meb_data(engine, 'nfi_meb', 'South')
        create_combined_dataframe(val, yoy, 'YoY').to_excel(writer, sheet_name='South YoY - NFI', index=False)
    
    # Apply formatting to YoY file
    print(f"\nApplying color formatting to YoY file...")
    apply_formatting(yoy_output)
    
    print(f"\n✓ YoY file saved: {yoy_output.name}")
    print(f"  Size: {yoy_output.stat().st_size / 1024:.1f} KB")
    
    # Final summary
    print(f"\n" + "="*70)
    print("EXPORT COMPLETE")
    print("="*70)
    print(f"\n✓ Two files created:")
    print(f"\n  1. MoM_MEB_TrendAnalysis.xlsx (12 sheets)")
    print(f"     Location: {mom_output}")
    print(f"     Size: {mom_output.stat().st_size / 1024:.1f} KB")
    print(f"     Sheets:")
    print(f"       - Master MoM (Full, Food, NFI)")
    print(f"       - East MoM (Full, Food, NFI)")
    print(f"       - West MoM (Full, Food, NFI)")
    print(f"       - South MoM (Full, Food, NFI)")

    print(f"\n  2. YoY_MEB_TrendAnalysis.xlsx (12 sheets)")
    print(f"     Location: {yoy_output}")
    print(f"     Size: {yoy_output.stat().st_size / 1024:.1f} KB")
    print(f"     Sheets:")
    print(f"       - Master YoY (Full, Food, NFI)")
    print(f"       - East YoY (Full, Food, NFI)")
    print(f"       - West YoY (Full, Food, NFI)")
    print(f"       - South YoY (Full, Food, NFI)")
    
    print(f"\n  Format:")
    print(f"    - Values: digits only (no 'LYD')")
    print(f"    - Changes: ▲ Red (increase) / ▼ Green (decrease)")
    
    return mom_output, yoy_output

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Export MEB Master sheets with MoM% and YoY% changes",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export MEB Trend Analysis with MoM% and YoY%
  python scripts/05_Data_Outputs/master_data.py

Output:
  TWO separate Excel files with date ranges in filenames:
  
  1. MoM_MEB_TrendAnalysis_YYYYMM-YYYYMM.xlsx (Month-over-Month changes)
     Example: MoM_MEB_TrendAnalysis_202304-202511.xlsx
     (12 sheets total)
     MASTER SHEETS (3 sheets):
       - Master MoM - Full, Food, Non-Food (All municipalities + regions)
     EAST REGION (3 sheets - 6 municipalities):
       - East MoM - Full, Food, Non-Food
     WEST REGION (3 sheets - 8 municipalities):
       - West MoM - Full, Food, Non-Food
     SOUTH REGION (3 sheets - 7 municipalities):
       - South MoM - Full, Food, Non-Food

  2. YoY_MEB_TrendAnalysis_YYYYMM-YYYYMM.xlsx (Year-over-Year changes)
     Example: YoY_MEB_TrendAnalysis_202304-202511.xlsx
     (12 sheets total)
  
  Each file contains:
     MASTER SHEETS (3 sheets):
       - Master YoY - Full, Food, Non-Food (All municipalities + regions)
     EAST REGION (3 sheets - 6 municipalities):
       - East YoY - Full, Food, Non-Food
     WEST REGION (3 sheets - 8 municipalities):
       - West YoY - Full, Food, Non-Food
     SOUTH REGION (3 sheets - 7 municipalities):
       - South YoY - Full, Food, Non-Food

Prerequisites:
  Database must be loaded with PMM data
        """
    )
    
    args = parser.parse_args()
    
    try:
        mom_file, yoy_file = export_historical_meb()
        print(f"\n✅ Ready to use in Excel, PowerPoint, or Dashboards!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)