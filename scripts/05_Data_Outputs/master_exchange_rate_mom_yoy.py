"""
Libya PMM - Master Exchange Rate MoM/YoY Analysis
Generates Month-over-Month and Year-over-Year analysis for exchange rates from database

Creates Excel file with 2 sheets:
  - MoM (Month-over-Month)
  - YoY (Year-over-Year)

Format: Alternating rows (like MEB files)
  - Row 1: Month + exchange rate values
  - Row 2: Empty + MoM%/YoY% changes with ▲/▼
  - Month cells are merged vertically

Filename: MoM_YoY_ExchangeRate_TrendAnalysis_YYYYMM-YYYYMM.xlsx
Location: Master Data/Exchange Rate/

Usage:
    python scripts/05_Data_Outputs/master_exchange_rate_mom_yoy.py
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import get_engine, PATHS

# ============================================================================
# EXCHANGE RATE OUTPUT DIRECTORY
# ============================================================================

EXCHANGE_RATE_DIR = PATHS['master_data'].parent / 'Exchange Rate'

# ============================================================================
# DATA EXTRACTION
# ============================================================================

def get_exchange_rate_data(engine):
    """
    Get all exchange rate data from database
    
    Returns DataFrame with columns:
    - date
    - official_rate
    - official_rate_with_tax
    - parallel_market_rate
    """
    
    query = """
    SELECT 
        date,
        official_rate,
        official_rate_with_tax,
        parallel_market_rate
    FROM exchange_rates
    ORDER BY date
    """
    
    print(f"\n📊 Querying exchange rate data from database...")
    df = pd.read_sql(query, engine)
    
    print(f"   ✓ Loaded {len(df)} daily records")
    print(f"   Date range: {df['date'].min()} to {df['date'].max()}")
    
    return df

# ============================================================================
# MONTHLY AGGREGATION
# ============================================================================

def calculate_monthly_averages(df):
    """
    Calculate monthly average exchange rates
    
    Returns DataFrame with one row per month
    """
    
    print(f"\n📅 Calculating monthly averages...")
    
    # Convert date to month period
    df['month'] = pd.to_datetime(df['date']).dt.to_period('M')
    
    # Group by month and calculate averages
    monthly = df.groupby('month').agg({
        'official_rate': 'mean',
        'official_rate_with_tax': 'mean',
        'parallel_market_rate': 'mean'
    }).reset_index()
    
    # Convert period back to timestamp for easier manipulation
    monthly['date'] = monthly['month'].dt.to_timestamp()
    
    # Round to 2 decimal places
    monthly['official_rate'] = monthly['official_rate'].round(2)
    monthly['official_rate_with_tax'] = monthly['official_rate_with_tax'].round(2)
    monthly['parallel_market_rate'] = monthly['parallel_market_rate'].round(2)
    
    print(f"   ✓ Calculated {len(monthly)} months")
    print(f"   Month range: {monthly['date'].min().strftime('%B %Y')} to {monthly['date'].max().strftime('%B %Y')}")
    
    return monthly

# ============================================================================
# MOM CALCULATION
# ============================================================================

def calculate_mom(monthly_df):
    """
    Calculate Month-over-Month changes
    
    Returns DataFrame with MoM values and percentages
    """
    
    print(f"\n📈 Calculating MoM changes...")
    
    df = monthly_df.copy()
    
    # Sort by date
    df = df.sort_values('date').reset_index(drop=True)
    
    # Calculate MoM percentage change
    df['official_mom_pct'] = ((df['official_rate'] - df['official_rate'].shift(1)) / df['official_rate'].shift(1) * 100)
    df['official_tax_mom_pct'] = ((df['official_rate_with_tax'] - df['official_rate_with_tax'].shift(1)) / df['official_rate_with_tax'].shift(1) * 100)
    df['parallel_mom_pct'] = ((df['parallel_market_rate'] - df['parallel_market_rate'].shift(1)) / df['parallel_market_rate'].shift(1) * 100)
    
    # Round percentages to 1 decimal place
    df['official_mom_pct'] = df['official_mom_pct'].round(1)
    df['official_tax_mom_pct'] = df['official_tax_mom_pct'].round(1)
    df['parallel_mom_pct'] = df['parallel_mom_pct'].round(1)
    
    print(f"   ✓ Calculated MoM for {len(df)} months")
    
    return df

# ============================================================================
# YOY CALCULATION
# ============================================================================

def calculate_yoy(monthly_df):
    """
    Calculate Year-over-Year changes
    
    Returns DataFrame with YoY values and percentages
    """
    
    print(f"\n📊 Calculating YoY changes...")
    
    df = monthly_df.copy()
    
    # Sort by date
    df = df.sort_values('date').reset_index(drop=True)
    
    # Calculate YoY percentage change
    df['official_yoy_pct'] = ((df['official_rate'] - df['official_rate'].shift(12)) / df['official_rate'].shift(12) * 100)
    df['official_tax_yoy_pct'] = ((df['official_rate_with_tax'] - df['official_rate_with_tax'].shift(12)) / df['official_rate_with_tax'].shift(12) * 100)
    df['parallel_yoy_pct'] = ((df['parallel_market_rate'] - df['parallel_market_rate'].shift(12)) / df['parallel_market_rate'].shift(12) * 100)
    
    # Round percentages to 1 decimal place
    df['official_yoy_pct'] = df['official_yoy_pct'].round(1)
    df['official_tax_yoy_pct'] = df['official_tax_yoy_pct'].round(1)
    df['parallel_yoy_pct'] = df['parallel_yoy_pct'].round(1)
    
    print(f"   ✓ Calculated YoY for {len(df)} months")
    
    return df

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def format_percentage(value):
    """Helper function to format percentage values with ▲/▼ indicators"""
    if pd.notna(value):
        if value > 0:
            return f"▲ +{value:.1f}%"
        elif value < 0:
            return f"▼ {value:.1f}%"
        else:
            return f"{value:.1f}%"
    else:
        return ""

# ============================================================================
# FORMAT FOR EXCEL
# ============================================================================

def format_mom_sheet(df):
    """
    Format MoM data for Excel output (alternating rows like MEB files)
    
    Structure for each month:
    Row 1: Month label in col A, values in cols B, C, D
    Row 2: Empty in col A, MoM% in cols B, C, D
    
    Columns:
    A: Month (Jan 2024)
    B: Official USD/LYD
    C: Official + Gov Tax
    D: Parallel Market
    """
    
    print(f"\n📄 Formatting MoM sheet...")
    
    # Create month labels (e.g., "Jan 2024")
    df['month_label'] = df['date'].dt.strftime('%b %Y')
    
    # Create alternating rows: value row, then % change row
    result_rows = []
    
    for idx, row in df.iterrows():
        # Value row
        value_row = [
            row['month_label'],
            f"{row['official_rate']:.2f}" if pd.notna(row['official_rate']) else "",
            f"{row['official_rate_with_tax']:.2f}" if pd.notna(row['official_rate_with_tax']) else "",
            f"{row['parallel_market_rate']:.2f}" if pd.notna(row['parallel_market_rate']) else ""
        ]
        result_rows.append(value_row)
        
        # Percentage change row
        pct_row = [
            "",  # Empty for month column
            format_percentage(row['official_mom_pct']),
            format_percentage(row['official_tax_mom_pct']),
            format_percentage(row['parallel_mom_pct'])
        ]
        result_rows.append(pct_row)
    
    # Create DataFrame with column names
    columns = ['Month', 'Official USD/LYD', 'Official + Gov Tax', 'Parallel Market']
    result_df = pd.DataFrame(result_rows, columns=columns)
    
    print(f"   ✓ Formatted {len(df)} months ({len(result_df)} rows with alternating structure)")
    
    return result_df

def format_yoy_sheet(df):
    """
    Format YoY data for Excel output (alternating rows like MEB files)
    
    Structure for each month:
    Row 1: Month label in col A, values in cols B, C, D
    Row 2: Empty in col A, YoY% in cols B, C, D
    
    Columns:
    A: Month (Jan 2024)
    B: Official USD/LYD
    C: Official + Gov Tax
    D: Parallel Market
    """
    
    print(f"\n📄 Formatting YoY sheet...")
    
    # Create month labels (e.g., "Jan 2024")
    df['month_label'] = df['date'].dt.strftime('%b %Y')
    
    # Create alternating rows: value row, then % change row
    result_rows = []
    
    for idx, row in df.iterrows():
        # Value row
        value_row = [
            row['month_label'],
            f"{row['official_rate']:.2f}" if pd.notna(row['official_rate']) else "",
            f"{row['official_rate_with_tax']:.2f}" if pd.notna(row['official_rate_with_tax']) else "",
            f"{row['parallel_market_rate']:.2f}" if pd.notna(row['parallel_market_rate']) else ""
        ]
        result_rows.append(value_row)
        
        # Percentage change row
        pct_row = [
            "",  # Empty for month column
            format_percentage(row['official_yoy_pct']),
            format_percentage(row['official_tax_yoy_pct']),
            format_percentage(row['parallel_yoy_pct'])
        ]
        result_rows.append(pct_row)
    
    # Create DataFrame with column names
    columns = ['Month', 'Official USD/LYD', 'Official + Gov Tax', 'Parallel Market']
    result_df = pd.DataFrame(result_rows, columns=columns)
    
    print(f"   ✓ Formatted {len(df)} months ({len(result_df)} rows with alternating structure)")
    
    return result_df

# ============================================================================
# APPLY FORMATTING
# ============================================================================

def apply_formatting(workbook_path):
    """Apply formatting: colors for ▲/▼, freeze panes, column widths, merge month cells"""
    
    print(f"\n🎨 Applying formatting...")
    
    wb = load_workbook(workbook_path)
    
    # Red and green fonts for percentage changes
    red_font = Font(name='Aptos Narrow', size=11, color='C00000')
    green_font = Font(name='Aptos Narrow', size=11, color='00B050')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Freeze panes at B2 (keep Month column and header row visible)
        ws.freeze_panes = 'B2'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Merge month cells (column A) for alternating rows
        # Every odd row (3, 5, 7, ...) after header should merge with the row above
        for row_idx in range(2, ws.max_row + 1, 2):  # Start at 2 (first data row), step by 2
            # Merge current row with next row in column A (Month column)
            if row_idx + 1 <= ws.max_row:
                ws.merge_cells(f'A{row_idx}:A{row_idx+1}')
                # Center align the merged cell
                ws.cell(row=row_idx, column=1).alignment = Alignment(vertical='center')
        
        # Apply red/green colors to percentage cells (odd rows after header, columns B-D)
        for row_idx in range(3, ws.max_row + 1, 2):  # Start at 3 (first % row), step by 2
            for col_idx in range(2, 5):  # Columns B, C, D (Official, Official+Tax, Parallel)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('▲'):
                        cell.font = red_font
                    elif cell.value.startswith('▼'):
                        cell.font = green_font
    
    wb.save(workbook_path)
    print(f"   ✓ Formatting applied: merged cells, colors, freeze panes")

# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def export_exchange_rate_analysis():
    """Export MoM and YoY exchange rate analysis to Excel"""
    
    engine = get_engine()
    
    print("="*70)
    print("LIBYA PMM - EXCHANGE RATE MoM/YoY ANALYSIS")
    print("="*70)
    
    # Create output directory if needed
    EXCHANGE_RATE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Get data
    raw_data = get_exchange_rate_data(engine)
    monthly_data = calculate_monthly_averages(raw_data)
    
    # Get date range for filename
    start_date = monthly_data['date'].min()
    end_date = monthly_data['date'].max()
    start_yyyymm = start_date.strftime('%Y%m')
    end_yyyymm = end_date.strftime('%Y%m')
    date_range_str = f"{start_yyyymm}-{end_yyyymm}"
    
    print(f"\n📅 Date range for filename: {date_range_str}")
    
    # Output file with date range in Exchange Rate directory
    output_file = EXCHANGE_RATE_DIR / f'MoM_YoY_ExchangeRate_TrendAnalysis_{date_range_str}.xlsx'
    
    # Clean up old files
    print(f"\n🗑️  Cleaning up old files...")
    old_files = list(EXCHANGE_RATE_DIR.glob('MoM_YoY_ExchangeRate_TrendAnalysis_*.xlsx'))
    
    deleted_count = 0
    for old_file in old_files:
        if old_file != output_file:
            print(f"   Deleting old file: {old_file.name}")
            old_file.unlink()
            deleted_count += 1
    
    if deleted_count > 0:
        print(f"   ✓ Deleted {deleted_count} old file(s)")
    else:
        print(f"   No old files to delete")
    
    # Calculate changes
    mom_data = calculate_mom(monthly_data)
    yoy_data = calculate_yoy(monthly_data)
    
    # Format sheets
    mom_sheet = format_mom_sheet(mom_data)
    yoy_sheet = format_yoy_sheet(yoy_data)
    
    # Write to Excel
    print(f"\n📝 Writing to Excel...")
    print(f"   File: {output_file.name}")
    print(f"   Location: {EXCHANGE_RATE_DIR}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        mom_sheet.to_excel(writer, sheet_name='MoM', index=False)
        yoy_sheet.to_excel(writer, sheet_name='YoY', index=False)
    
    print(f"   ✓ Sheets written: MoM, YoY")
    
    # Apply formatting
    apply_formatting(output_file)
    
    print(f"\n" + "="*70)
    print("EXPORT COMPLETE")
    print("="*70)
    print(f"\n✓ File saved: {output_file.name}")
    print(f"  Location: {output_file}")
    print(f"  Size: {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  Format: Alternating rows (value, then %) like MEB files")
    print(f"  Sheets: MoM (Month-over-Month), YoY (Year-over-Year)")
    
    return output_file

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Export Exchange Rate MoM/YoY Analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export exchange rate analysis
  python scripts/05_Data_Outputs/master_exchange_rate_mom_yoy.py

Output:
  MoM_YoY_ExchangeRate_TrendAnalysis_YYYYMM-YYYYMM.xlsx
  Example: MoM_YoY_ExchangeRate_TrendAnalysis_202301-202511.xlsx
  Location: Master Data/Exchange Rate/
  
  Two sheets:
  - MoM (Month-over-Month % changes)
  - YoY (Year-over-Year % changes)
  
  Format (alternating rows like MEB files):
  ┌─────────────┬──────────────────┬──────────────────┬─────────────────┐
  │ Month       │ Official USD/LYD │ Official + Tax   │ Parallel Market │
  ├─────────────┼──────────────────┼──────────────────┼─────────────────┤
  │ Jan 2024    │ 4.85             │ 5.82             │ 7.20            │
  │             │ ▲ +2.1%          │ ▲ +2.1%          │ ▲ +1.5%         │
  │ Feb 2024    │ 4.90             │ 5.88             │ 7.30            │
  │             │ ▲ +1.0%          │ ▲ +1.0%          │ ▲ +1.4%         │
  └─────────────┴──────────────────┴──────────────────┴─────────────────┘
  
  - Colors: Red (▲ increase), Green (▼ decrease)
  - Freeze panes at B2 (Month column + header row frozen)
  - Month cells merged vertically
  - Auto-adjusted column widths

Features:
  - Dynamic date range in filename
  - Auto-deletes old files
  - Matches MEB file structure
  - No "nan" values in output

Prerequisites:
  Database must be loaded with exchange rate data
        """
    )
    
    args = parser.parse_args()
    
    try:
        output_file = export_exchange_rate_analysis()
        print(f"\n✅ Ready to use in Excel or PowerPoint!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)