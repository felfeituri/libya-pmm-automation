"""
Libya PMM - Commodity Price Comparison Tables
Generates Excel tables showing current vs previous month prices with MoM% change

Format:
- Product name in merged cell spanning 2 rows
- Row 1: Prices for both months
- Row 2: MoM% for both months
- Previous month MoM% compares to month before it
- Current month MoM% compares to previous month

Usage:
    python scripts/05_Data_Outputs/commodity_tables.py <year> <month>

Examples:
    python scripts/05_Data_Outputs/commodity_tables.py 2025 11
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from datetime import datetime
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import get_month_paths, ensure_month_directories, get_engine

# ============================================================================
# PRODUCT ORDERING
# ============================================================================

FOOD_PRODUCTS = [
    'Bread (5pc)',
    'Rice (Kg)',
    'Couscous (Kg)',
    'Pasta (500g)',
    'Beans (400g)',
    'Potatoes (Kg)',
    'Tomatoes (Kg)',
    'Pepper (Kg)',
    'Onions (Kg)',
    'Tomato Paste (400g)',
    'Chicken (Kg)',
    'Eggs (30pc)',
    'Tuna (200g)',
    'Milk (L)',
    'Oil (L)',
    'Salt (Kg)',
    'Sugar (Kg)',
    'Black Tea (250g)'
]

NFI_PRODUCTS = [
    'Handwash Soap (Pc)',
    'Dishwashing Liquid (L)',
    'Laundry Detergent (L)',
    'Toothpaste (Pc)',
    'Sanitary Pads (10Pc)'
]

FUEL_PRODUCTS = [
    'Cooking Fuel (11Kg)'
]

# Display name mapping (database name -> display name)
PRODUCT_DISPLAY_NAMES = {
    'Cooking Fuel (11Kg)': 'Public Cooking Fuel (11Kg)'
}

# ============================================================================
# DATA EXTRACTION
# ============================================================================

def get_commodity_comparison(engine, target_date):
    """
    Get commodity prices for three months:
    - prev_prev (month before previous) - for calculating previous month MoM%
    - previous (reference month)
    - current (latest month)
    """
    
    target_dt = pd.to_datetime(target_date)
    current_month_start = target_dt.replace(day=1)
    
    # Previous month
    if current_month_start.month == 1:
        prev_month_start = current_month_start.replace(year=current_month_start.year - 1, month=12)
    else:
        prev_month_start = current_month_start.replace(month=current_month_start.month - 1)
    
    # Month before previous
    if prev_month_start.month == 1:
        prev_prev_month_start = prev_month_start.replace(year=prev_month_start.year - 1, month=12)
    else:
        prev_prev_month_start = prev_month_start.replace(month=prev_month_start.month - 1)
    
    query = text("""
        SELECT 
            date,
            product_name,
            admin_name,
            average_price
        FROM products
        WHERE (date = :current_month OR date = :prev_month OR date = :prev_prev_month)
        AND average_price IS NOT NULL
        AND average_price > 0
        ORDER BY admin_name, product_name, date
    """)
    
    with engine.connect() as conn:
        result = conn.execute(query, {
            'current_month': current_month_start,
            'prev_month': prev_month_start,
            'prev_prev_month': prev_prev_month_start
        })
        df = pd.DataFrame(result.fetchall(), 
                         columns=['date', 'product_name', 'admin_name', 'average_price'])
    
    df['date'] = pd.to_datetime(df['date'])
    df['month_label'] = df['date'].dt.strftime('%B %Y')  # Full month name
    
    # Build result structure
    result_data = {}
    
    for admin in ['Libya', 'East', 'West', 'South']:
        admin_label = 'National' if admin == 'Libya' else admin
        admin_data = df[df['admin_name'] == admin]
        
        result_data[admin_label] = {}
        
        for product in FOOD_PRODUCTS + NFI_PRODUCTS + FUEL_PRODUCTS:
            product_data = admin_data[admin_data['product_name'] == product].sort_values('date')
            
            prev_prev_price = None
            prev_price = None
            curr_price = None
            prev_mom = None
            curr_mom = None
            
            if len(product_data) >= 3:
                # Have all three months
                prev_prev_price = float(product_data.iloc[0]['average_price'])
                prev_price = float(product_data.iloc[1]['average_price'])
                curr_price = float(product_data.iloc[2]['average_price'])
                
                # Previous month MoM: compare prev to prev_prev
                prev_mom = ((prev_price - prev_prev_price) / prev_prev_price) * 100
                
                # Current month MoM: compare curr to prev
                curr_mom = ((curr_price - prev_price) / prev_price) * 100
                
            elif len(product_data) == 2:
                # Have current and previous (no prev_prev)
                prev_price = float(product_data.iloc[0]['average_price'])
                curr_price = float(product_data.iloc[1]['average_price'])
                
                prev_mom = None  # Can't calculate
                curr_mom = ((curr_price - prev_price) / prev_price) * 100
                
            elif len(product_data) == 1:
                curr_price = float(product_data.iloc[0]['average_price'])
            
            result_data[admin_label][product] = {
                'current': curr_price,
                'previous': prev_price,
                'current_mom': curr_mom,
                'prev_mom': prev_mom
            }
    
    return result_data, prev_month_start.strftime('%B %Y'), current_month_start.strftime('%B %Y')

# ============================================================================
# EXCEL TABLE GENERATION
# ============================================================================

def add_commodity_sheet(wb, sheet_name, region_data, previous_month, current_month):
    """Add a commodity table sheet to existing workbook"""
    
    # Create new sheet or use active
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Styles - All 9pt for commodity tables
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    header_font = Font(name='Aptos Narrow', size=9, bold=True, color='595959')
    
    section_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    section_font = Font(name='Aptos Narrow', size=9, bold=True)
    
    product_font = Font(name='Aptos Narrow', size=9, bold=True, color='595959')
    value_font = Font(name='Aptos Narrow', size=9)
    red_font = Font(name='Aptos Narrow', size=9, color='C00000', bold=True)
    green_font = Font(name='Aptos Narrow', size=9, color='00B050', bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    value_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    mom_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    product_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    current_row = 1
    
    # Header row
    ws.cell(row=current_row, column=1, value='')
    ws.cell(row=current_row, column=2, value=previous_month)
    ws.cell(row=current_row, column=3, value=current_month)
    
    for col in range(1, 4):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    
    # Process each product category
    categories = [
        ('Food Items', FOOD_PRODUCTS),
        ('Non-Food Items', NFI_PRODUCTS),
        ('Fuel', FUEL_PRODUCTS)
    ]
    
    # Use the passed region_data directly
    for category_name, products in categories:
        # Category header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        category_cell = ws.cell(row=current_row, column=1, value=category_name)
        category_cell.font = section_font
        category_cell.fill = section_fill
        category_cell.alignment = left_align
        category_cell.border = thin_border
        current_row += 1
        
        # Products in category
        for product in products:
            product_info = region_data.get(product, {})
            display_name = PRODUCT_DISPLAY_NAMES.get(product, product)
            
            # Merge product name across 2 rows
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
            product_cell = ws.cell(row=current_row, column=1, value=display_name)
            product_cell.font = product_font
            product_cell.alignment = left_align
            product_cell.border = product_border
            
            # Row 1: Prices
            prev_price = product_info.get('previous')
            curr_price = product_info.get('current')
            
            prev_cell = ws.cell(row=current_row, column=2, 
                              value=f"LYD {prev_price:.2f}" if prev_price else "-")
            prev_cell.font = value_font
            prev_cell.alignment = center_align
            prev_cell.border = value_row_border
            
            curr_cell = ws.cell(row=current_row, column=3, 
                              value=f"LYD {curr_price:.2f}" if curr_price else "-")
            curr_cell.font = value_font
            curr_cell.alignment = center_align
            curr_cell.border = value_row_border
            
            current_row += 1
            
            # Row 2: MoM%
            prev_mom = product_info.get('prev_mom')
            curr_mom = product_info.get('current_mom')
            
            # Previous month MoM
            if prev_mom is not None:
                if prev_mom > 0:
                    prev_mom_str = f"▲ +{prev_mom:.1f}%"
                    prev_mom_font = red_font
                elif prev_mom < 0:
                    prev_mom_str = f"▼ {prev_mom:.1f}%"
                    prev_mom_font = green_font
                else:
                    prev_mom_str = f"{prev_mom:.1f}%"
                    prev_mom_font = value_font
            else:
                prev_mom_str = ""
                prev_mom_font = value_font
            
            prev_mom_cell = ws.cell(row=current_row, column=2, value=prev_mom_str)
            prev_mom_cell.font = prev_mom_font
            prev_mom_cell.alignment = center_align
            prev_mom_cell.border = mom_row_border
            
            # Current month MoM
            if curr_mom is not None:
                if curr_mom > 0:
                    curr_mom_str = f"▲ +{curr_mom:.1f}%"
                    curr_mom_font = red_font
                elif curr_mom < 0:
                    curr_mom_str = f"▼ {curr_mom:.1f}%"
                    curr_mom_font = green_font
                else:
                    curr_mom_str = f"{curr_mom:.1f}%"
                    curr_mom_font = value_font
            else:
                curr_mom_str = ""
                curr_mom_font = value_font
            
            curr_mom_cell = ws.cell(row=current_row, column=3, value=curr_mom_str)
            curr_mom_cell.font = curr_mom_font
            curr_mom_cell.alignment = center_align
            curr_mom_cell.border = mom_row_border
            
            # Border for second row of merged product cell
            ws.cell(row=current_row, column=1).border = product_border
            
            current_row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    
    # Row heights
    for row in range(1, current_row):
        ws.row_dimensions[row].height = 20
    
    # Freeze header
    ws.freeze_panes = 'A2'

# ============================================================================
# MAIN
# ============================================================================

def generate_commodity_tables(year, month):
    """Generate commodity comparison tables for specific month"""
    
    engine = get_engine()
    
    print("="*70)
    print("LIBYA PMM - COMMODITY PRICE COMPARISON TABLES")
    print("="*70)
    
    # Get paths from config
    paths = ensure_month_directories(year, month)
    output_file = paths['commodity_comparison']
    
    print(f"\nGenerating tables for: {paths['month_name']} {year}")
    print(f"Month tag: {paths['month_tag']}")
    
    # Get data using the target month date
    target_date = paths['date']
    
    print(f"\nExtracting commodity price data...")
    data, previous_month, current_month = get_commodity_comparison(engine, target_date)
    
    print(f"Previous month: {previous_month}")
    print(f"Current month: {current_month}")
    
    # Create workbook with all regional sheets
    print(f"\nCreating Excel tables...")
    wb = Workbook()
    
    # Remove default sheet if needed
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Add sheets for each region in order
    regions = ['National', 'East', 'West', 'South']
    for i, region in enumerate(regions):
        print(f"  {i+1}. Creating {region} sheet...")
        if region in data:
            add_commodity_sheet(wb, region, data[region], previous_month, current_month)
        else:
            print(f"     ⚠ Warning: No data found for {region}")
    
    # Save workbook
    wb.save(output_file)
    
    print("\n" + "="*70)
    print("TABLE GENERATION COMPLETE")
    print("="*70)
    print(f"\n✓ Workbook saved: {output_file.name}")
    print(f"  Location: {paths['tables']}")
    print(f"  Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
    print(f"  Products: {len(FOOD_PRODUCTS)} Food + {len(NFI_PRODUCTS)} NFI + {len(FUEL_PRODUCTS)} Fuel")
    print(f"  File size: {output_file.stat().st_size / 1024:.1f} KB")
    
    return output_file

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generate commodity price comparison tables",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate tables for November 2025
  python scripts/05_Data_Outputs/commodity_tables.py 2025 11
  
  # Generate tables for October 2025
  python scripts/05_Data_Outputs/commodity_tables.py 2025 10

Output:
  Commodity_Price_Comparison_Nov25.xlsx with:
  - Current month vs previous month prices
  - MoM% changes for both months
  - Food Items, Non-Food Items, and Fuel sections

Prerequisites:
  Database must be loaded with at least 2 months of PMM data
        """
    )
    
    parser.add_argument('year', type=int, help='Year (e.g., 2025)')
    parser.add_argument('month', type=int, help='Month (1-12)')
    
    args = parser.parse_args()
    
    # Validate month
    if not 1 <= args.month <= 12:
        print(f"Error: Month must be between 1 and 12, got {args.month}")
        sys.exit(1)
    
    try:
        output_file = generate_commodity_tables(args.year, args.month)
        print(f"\n✅ Ready to copy into PowerPoint!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)