"""
Libya PMM - MEB Comparison Tables (Last 3 Months)
Generates Excel tables showing Full, Food, and NFI MEB for National, Regional, and Municipal levels

Format:
- 12 sheets total (3 MEB types × 4 geographic levels)
- National: National, East, West, South
- Mantika (All Municipalities): All 21 municipalities
- East: East region municipalities only
- West: West region municipalities only
- South: South region municipalities only
- Last 3 months in columns
- Region/Municipality name in merged cell spanning 2 rows
- Row 1: MEB values for each month
- Row 2: MoM% for each month
- WFP branding colors

Usage:
    python scripts/05_Data_Outputs/meb_tables.py <year> <month>

Examples:
    python scripts/05_Data_Outputs/meb_tables.py 2025 11
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from datetime import datetime
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import get_month_paths, ensure_month_directories, get_engine

# ============================================================================
# MUNICIPALITY ORDERING
# ============================================================================

# All municipalities in alphabetical order
ALL_MUNICIPALITIES = [
    'AlBayda', 'Algatroun', 'AlJufra', 'AlKhums', 'AlKufra', 'Azzawya',
    'Benghazi', 'Derna', 'Ejdabia', 'Ghat', 'Misrata', 'Murzuq', 'Nalut',
    'Sebha', 'Sirt', 'Tobruk', 'Tripoli Center', 'Ubari', 'Wadi Alshati',
    'Zliten', 'Zwara'
]

# East region municipalities
EAST_MUNICIPALITIES = ['AlBayda', 'Benghazi', 'Derna', 'Ejdabia', 'AlKufra', 'Tobruk']

# West region municipalities  
WEST_MUNICIPALITIES = ['AlKhums', 'Azzawya', 'Misrata', 'Nalut', 'Sirt', 'Tripoli Center', 'Zliten', 'Zwara']

# South region municipalities
SOUTH_MUNICIPALITIES = ['Algatroun', 'AlJufra', 'Ghat', 'Murzuq', 'Sebha', 'Ubari', 'Wadi Alshati']

# ============================================================================
# DATA EXTRACTION - NATIONAL
# ============================================================================

def get_meb_comparison(engine, target_date):
    """
    Get MEB data for last 3 months plus month before for MoM calculation
    Returns: data dict, month labels list
    """
    
    target_dt = pd.to_datetime(target_date)
    month3 = target_dt.replace(day=1)  # Most recent (e.g., November)
    
    # Calculate previous months
    if month3.month == 1:
        month2 = month3.replace(year=month3.year - 1, month=12)
    else:
        month2 = month3.replace(month=month3.month - 1)
    
    if month2.month == 1:
        month1 = month2.replace(year=month2.year - 1, month=12)
    else:
        month1 = month2.replace(month=month2.month - 1)
    
    # Month before month1 for MoM calculation
    if month1.month == 1:
        month0 = month1.replace(year=month1.year - 1, month=12)
    else:
        month0 = month1.replace(month=month1.month - 1)
    
    month_labels = [
        month1.strftime('%B %Y'),
        month2.strftime('%B %Y'),
        month3.strftime('%B %Y')
    ]
    
    # Get National data
    national_query = text("""
        SELECT 
            date,
            full_meb,
            food_meb,
            nfi_meb
        FROM national_meb
        WHERE date IN (:m0, :m1, :m2, :m3)
        ORDER BY date
    """)
    
    # Get Regional data
    regional_query = text("""
        SELECT 
            date,
            region,
            full_meb,
            food_meb,
            nfi_meb
        FROM regional_meb
        WHERE date IN (:m0, :m1, :m2, :m3)
        AND region IN ('East', 'West', 'South')
        ORDER BY date, region
    """)
    
    result_data = {}
    
    with engine.connect() as conn:
        # National
        nat_result = conn.execute(national_query, {
            'm0': month0, 'm1': month1, 'm2': month2, 'm3': month3
        })
        nat_df = pd.DataFrame(nat_result.fetchall(), 
                             columns=['date', 'full_meb', 'food_meb', 'nfi_meb'])
        
        nat_df['date'] = pd.to_datetime(nat_df['date'])
        nat_df = nat_df.sort_values('date')
        
        result_data['National'] = process_meb_data(nat_df, [month0, month1, month2, month3])
        
        # Regional
        reg_result = conn.execute(regional_query, {
            'm0': month0, 'm1': month1, 'm2': month2, 'm3': month3
        })
        reg_df = pd.DataFrame(reg_result.fetchall(), 
                             columns=['date', 'region', 'full_meb', 'food_meb', 'nfi_meb'])
        
        reg_df['date'] = pd.to_datetime(reg_df['date'])
        
        for region in ['East', 'West', 'South']:
            region_data = reg_df[reg_df['region'] == region].sort_values('date')
            result_data[region] = process_meb_data(region_data, [month0, month1, month2, month3])
    
    return result_data, month_labels

# ============================================================================
# DATA EXTRACTION - MUNICIPALITIES
# ============================================================================

def get_municipality_meb_comparison(engine, target_date, municipality_list):
    """
    Get municipality MEB data for last 3 months plus month before for MoM calculation
    Returns: data dict, month labels list
    """
    
    target_dt = pd.to_datetime(target_date)
    month3 = target_dt.replace(day=1)
    
    if month3.month == 1:
        month2 = month3.replace(year=month3.year - 1, month=12)
    else:
        month2 = month3.replace(month=month3.month - 1)
    
    if month2.month == 1:
        month1 = month2.replace(year=month2.year - 1, month=12)
    else:
        month1 = month2.replace(month=month2.month - 1)
    
    if month1.month == 1:
        month0 = month1.replace(year=month1.year - 1, month=12)
    else:
        month0 = month1.replace(month=month1.month - 1)
    
    month_labels = [
        month1.strftime('%B %Y'),
        month2.strftime('%B %Y'),
        month3.strftime('%B %Y')
    ]
    
    # Get Municipality data
    municipality_query = text("""
        SELECT 
            date,
            municipality,
            full_meb,
            food_meb,
            nfi_meb
        FROM municipality_meb
        WHERE date IN (:m0, :m1, :m2, :m3)
        AND municipality = ANY(:municipalities)
        ORDER BY date, municipality
    """)
    
    result_data = {}
    
    with engine.connect() as conn:
        muni_result = conn.execute(municipality_query, {
            'm0': month0, 'm1': month1, 'm2': month2, 'm3': month3,
            'municipalities': municipality_list
        })
        muni_df = pd.DataFrame(muni_result.fetchall(), 
                              columns=['date', 'municipality', 'full_meb', 'food_meb', 'nfi_meb'])
        
        muni_df['date'] = pd.to_datetime(muni_df['date'])
        
        for municipality in municipality_list:
            muni_data = muni_df[muni_df['municipality'] == municipality].sort_values('date')
            result_data[municipality] = process_meb_data(muni_data, [month0, month1, month2, month3])
    
    return result_data, month_labels

# ============================================================================
# DATA PROCESSING
# ============================================================================

def process_meb_data(df, months):
    """Process MEB data and calculate MoM%"""
    result = {
        'full_meb': [],
        'food_meb': [],
        'nfi_meb': [],
        'full_meb_mom': [],
        'food_meb_mom': [],
        'nfi_meb_mom': []
    }
    
    # Extract values for each month (months[0] is for MoM calc only)
    values = {
        'full_meb': [],
        'food_meb': [],
        'nfi_meb': []
    }
    
    for month in months:
        month_data = df[df['date'] == month]
        if len(month_data) > 0:
            values['full_meb'].append(float(month_data.iloc[0]['full_meb']) if month_data.iloc[0]['full_meb'] else None)
            values['food_meb'].append(float(month_data.iloc[0]['food_meb']) if month_data.iloc[0]['food_meb'] else None)
            values['nfi_meb'].append(float(month_data.iloc[0]['nfi_meb']) if month_data.iloc[0]['nfi_meb'] else None)
        else:
            values['full_meb'].append(None)
            values['food_meb'].append(None)
            values['nfi_meb'].append(None)
    
    # Store last 3 months values and calculate MoM
    for meb_type in ['full_meb', 'food_meb', 'nfi_meb']:
        for i in range(1, 4):  # months 1, 2, 3 (skip month 0 which is just for calc)
            current = values[meb_type][i]
            previous = values[meb_type][i-1]
            
            result[meb_type].append(current)
            
            if current is not None and previous is not None and previous > 0:
                mom = ((current - previous) / previous) * 100
                result[f'{meb_type}_mom'].append(mom)
            else:
                result[f'{meb_type}_mom'].append(None)
    
    return result

# ============================================================================
# EXCEL TABLE GENERATION
# ============================================================================

def create_combined_meb_sheet(ws, sheet_name, data, month_labels, entity_list):
    """Create a sheet with Full MEB, Food MEB, and NFI MEB all combined"""
    
    # Styles
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    header_font = Font(name='Aptos Narrow', size=11, bold=True, color='595959')
    
    section_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    section_font = Font(name='Aptos Narrow', size=12, bold=True, color='000000')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    value_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    mom_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    region_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Font sizes - 12pt for National, 9pt for others
    if 'National' in sheet_name:
        region_font_size = Font(name='Aptos Narrow', size=12, bold=True, color='595959')
        value_font_size = Font(name='Aptos Narrow', size=12)
        red_font_size = Font(name='Aptos Narrow', size=12, color='C00000', bold=True)
        green_font_size = Font(name='Aptos Narrow', size=12, color='00B050', bold=True)
    else:
        region_font_size = Font(name='Aptos Narrow', size=9, bold=True, color='595959')
        value_font_size = Font(name='Aptos Narrow', size=9)
        red_font_size = Font(name='Aptos Narrow', size=9, color='C00000', bold=True)
        green_font_size = Font(name='Aptos Narrow', size=9, color='00B050', bold=True)
    
    current_row = 1
    
    # Header row
    ws.cell(row=current_row, column=1, value='')
    for col_idx, month_label in enumerate(month_labels, start=2):
        ws.cell(row=current_row, column=col_idx, value=month_label)
    
    for col in range(1, len(month_labels) + 2):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    
    # Loop through all three MEB types
    for meb_type, meb_label in [('full_meb', 'Full MEB'), ('food_meb', 'Food MEB'), ('nfi_meb', 'NFI MEB')]:
        
        # Section header (Full MEB, Food MEB, or NFI MEB)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(month_labels) + 1)
        section_cell = ws.cell(row=current_row, column=1, value=meb_label)
        section_cell.font = section_font
        section_cell.alignment = center_align
        section_cell.fill = section_fill
        section_cell.border = thin_border
        
        current_row += 1
        
        # Data rows for each entity
        for entity in entity_list:
            if entity not in data:
                continue
                
            entity_data = data[entity]
            
            # Entity name (merged across 2 rows)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
            entity_cell = ws.cell(row=current_row, column=1, value=entity)
            entity_cell.font = region_font_size
            entity_cell.alignment = left_align
            entity_cell.border = region_border
            
            # Row 1: MEB values
            for col_idx in range(3):
                value = entity_data[meb_type][col_idx]
                value_str = f"LYD {value:.2f}" if value is not None else "-"
                
                cell = ws.cell(row=current_row, column=col_idx + 2, value=value_str)
                cell.font = value_font_size
                cell.alignment = center_align
                cell.border = value_row_border
            
            current_row += 1
            
            # Row 2: MoM%
            for col_idx in range(3):
                mom = entity_data[f'{meb_type}_mom'][col_idx]
                
                if mom is not None:
                    if mom > 0:
                        mom_str = f"▲ +{mom:.1f}%"
                        mom_font = red_font_size
                    elif mom < 0:
                        mom_str = f"▼ {mom:.1f}%"
                        mom_font = green_font_size
                    else:
                        mom_str = f"{mom:.1f}%"
                        mom_font = value_font_size
                else:
                    mom_str = ""
                    mom_font = value_font_size
                
                cell = ws.cell(row=current_row, column=col_idx + 2, value=mom_str)
                cell.font = mom_font
                cell.alignment = center_align
                cell.border = mom_row_border
            
            # Border for second row of merged entity cell
            ws.cell(row=current_row, column=1).border = region_border
            
            current_row += 1
        
        # No extra space between sections - removed the current_row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(month_labels) + 2):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Row heights
    for row in range(1, current_row):
        ws.row_dimensions[row].height = 20
    
    # Freeze header
    ws.freeze_panes = 'A2'

def create_meb_sheet(ws, sheet_name, data, month_labels, entity_list, meb_field):
    """Create a single MEB sheet with proper formatting"""
    
    # Styles
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    header_font = Font(name='Aptos Narrow', size=11, bold=True, color='595959')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    value_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    mom_row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    region_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Font sizes based on sheet type
    # Master sheets (National) = 12pt
    # All other sheets (Mantika, East, West, South) = 9pt
    if 'National' in sheet_name:
        # Master sheets: 12pt for everything
        region_font_size = Font(name='Aptos Narrow', size=12, bold=True, color='595959')
        value_font_size = Font(name='Aptos Narrow', size=12)
        red_font_size = Font(name='Aptos Narrow', size=12, color='C00000', bold=True)
        green_font_size = Font(name='Aptos Narrow', size=12, color='00B050', bold=True)
    else:
        # All other sheets: 9pt for everything
        region_font_size = Font(name='Aptos Narrow', size=9, bold=True, color='595959')
        value_font_size = Font(name='Aptos Narrow', size=9)
        red_font_size = Font(name='Aptos Narrow', size=9, color='C00000', bold=True)
        green_font_size = Font(name='Aptos Narrow', size=9, color='00B050', bold=True)
    
    current_row = 1
    
    # Header row
    ws.cell(row=current_row, column=1, value='')
    for col_idx, month_label in enumerate(month_labels, start=2):
        ws.cell(row=current_row, column=col_idx, value=month_label)
    
    for col in range(1, len(month_labels) + 2):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += 1
    
    # Data rows for each entity (region or municipality)
    for entity in entity_list:
        if entity not in data:
            continue
            
        entity_data = data[entity]
        
        # Entity name (merged across 2 rows)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
        entity_cell = ws.cell(row=current_row, column=1, value=entity)
        entity_cell.font = region_font_size
        entity_cell.alignment = left_align
        entity_cell.border = region_border
        
        # Row 1: MEB values
        for col_idx in range(3):
            value = entity_data[meb_field][col_idx]
            value_str = f"LYD {value:.2f}" if value is not None else "-"
            
            cell = ws.cell(row=current_row, column=col_idx + 2, value=value_str)
            cell.font = value_font_size
            cell.alignment = center_align
            cell.border = value_row_border
        
        current_row += 1
        
        # Row 2: MoM%
        for col_idx in range(3):
            mom = entity_data[f'{meb_field}_mom'][col_idx]
            
            if mom is not None:
                if mom > 0:
                    mom_str = f"▲ +{mom:.1f}%"
                    mom_font = red_font_size
                elif mom < 0:
                    mom_str = f"▼ {mom:.1f}%"
                    mom_font = green_font_size
                else:
                    mom_str = f"{mom:.1f}%"
                    mom_font = value_font_size
            else:
                mom_str = ""
                mom_font = value_font_size
            
            cell = ws.cell(row=current_row, column=col_idx + 2, value=mom_str)
            cell.font = mom_font
            cell.alignment = center_align
            cell.border = mom_row_border
        
        # Border for second row of merged entity cell
        ws.cell(row=current_row, column=1).border = region_border
        
        current_row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(month_labels) + 2):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Row heights
    for row in range(1, current_row):
        ws.row_dimensions[row].height = 20
    
    # Freeze header
    ws.freeze_panes = 'A2'

def create_meb_tables_excel(nat_reg_data, mantika_data, east_data, west_data, south_data, month_labels, output_file):
    """Create formatted Excel tables with all MEB types combined per geographic level"""
    
    wb = Workbook()
    
    # Each sheet contains Full MEB, Food MEB, and NFI MEB for one geographic level
    # Removed Mantika sheet per user request
    sheets_config = [
        ('National', nat_reg_data, ['National', 'East', 'West', 'South']),
        ('East', east_data, EAST_MUNICIPALITIES),
        ('West', west_data, WEST_MUNICIPALITIES),
        ('South', south_data, SOUTH_MUNICIPALITIES),
    ]
    
    for idx, (sheet_name, data, entity_list) in enumerate(sheets_config):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Create combined sheet with all three MEB types
        create_combined_meb_sheet(ws, sheet_name, data, month_labels, entity_list)
    
    wb.save(output_file)

# ============================================================================
# MAIN
# ============================================================================

def generate_meb_tables(year, month):
    """Generate MEB comparison tables for specific month"""
    
    engine = get_engine()
    
    print("="*70)
    print("LIBYA PMM - MEB COMPARISON TABLES (LAST 3 MONTHS)")
    print("="*70)
    
    # Get paths from config
    paths = ensure_month_directories(year, month)
    output_file = paths['meb_comparison']
    
    print(f"\nGenerating tables for: {paths['month_name']} {year}")
    print(f"Month tag: {paths['month_tag']}")
    
    # Get data using the target month date
    target_date = paths['date']
    
    print(f"\nExtracting data for last 3 months...")
    
    # National
    print(f"  - National data...")
    nat_reg_data, month_labels = get_meb_comparison(engine, target_date)
    
    # Mantika (All Municipalities)
    print(f"  - All municipalities (Mantika)...")
    mantika_data, _ = get_municipality_meb_comparison(engine, target_date, ALL_MUNICIPALITIES)
    
    # East
    print(f"  - East municipalities...")
    east_data, _ = get_municipality_meb_comparison(engine, target_date, EAST_MUNICIPALITIES)
    
    # West
    print(f"  - West municipalities...")
    west_data, _ = get_municipality_meb_comparison(engine, target_date, WEST_MUNICIPALITIES)
    
    # South
    print(f"  - South municipalities...")
    south_data, _ = get_municipality_meb_comparison(engine, target_date, SOUTH_MUNICIPALITIES)
    
    print(f"\nMonths: {', '.join(month_labels)}")
    
    print(f"\nCreating Excel workbook with 15 sheets...")
    create_meb_tables_excel(nat_reg_data, mantika_data, east_data, west_data, south_data, 
                           month_labels, output_file)
    
    print("\n" + "="*70)
    print("TABLE GENERATION COMPLETE")
    print("="*70)
    print(f"\n✓ Workbook saved: {output_file.name}")
    print(f"  Location: {paths['tables']}")
    print(f"  4 sheets (each with Full, Food, and NFI MEB):")
    print(f"    - National")
    print(f"    - East ({len(EAST_MUNICIPALITIES)} municipalities)")
    print(f"    - West ({len(WEST_MUNICIPALITIES)} municipalities)")
    print(f"    - South ({len(SOUTH_MUNICIPALITIES)} municipalities)")
    print(f"  File size: {output_file.stat().st_size / 1024:.1f} KB")
    
    return output_file

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generate MEB comparison tables for last 3 months",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate tables for November 2025
  python scripts/04_Data_Outputs/meb_tables.py 2025 11
  
  # Generate tables for October 2025
  python scripts/04_Data_Outputs/meb_tables.py 2025 10

Output:
  MEB_Comparison_Nov25.xlsx with 4 sheets:
  
  Each sheet contains Full MEB, Food MEB, and NFI MEB:
  
  1. National
     - National, East, West, South
  
  2. East
     - 6 municipalities (AlBayda, Benghazi, Derna, Ejdabia, AlKufra, Tobruk)
  
  3. West
     - 8 municipalities (AlKhums, Azzawya, Misrata, Nalut, Sirt, Tripoli Center, Zliten, Zwara)
  
  4. South
     - 7 municipalities (Algatroun, AlJufra, Ghat, Murzuq, Sebha, Ubari, Wadi Alshati)

Prerequisites:
  Database must be loaded with at least 3 months of PMM data
        """
    )
    
    parser.add_argument('year', type=int, help='Year (e.g., 2025)')
    parser.add_argument('month', type=int, help='Month (1-12)')
    
    args = parser.parse_args()
    
    # Validate month
    if not 1 <= args.month <= 12:
        print(f"Error: Month must be between 1 and 12, got {args.month}")
        sys.exit(1)
    
    try:
        output_file = generate_meb_tables(args.year, args.month)
        print(f"\n✅ Ready to copy into PowerPoint!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)