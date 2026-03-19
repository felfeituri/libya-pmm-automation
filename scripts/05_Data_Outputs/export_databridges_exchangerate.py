"""
Libya PMM - Export DataBridges Exchange Rate
Appends monthly exchange rate averages to DataBridges Exchange Rate file

This script:
1. Reads monthly exchange rate file from OneDrive
2. Calculates average official and parallel rates for the month
3. Appends or updates the row in the DataBridges Exchange Rate file
4. If the month already exists, it is overwritten (non-interactive)

Usage:
    python scripts/05_Data_Outputs/export_databridges_exchangerate.py <year> <month>

Examples:
    python scripts/05_Data_Outputs/export_databridges_exchangerate.py 2025 11
"""

import sys
from pathlib import Path
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/05_Data_Outputs/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
from config import get_month_paths, DATABRIDGES_BASE

# ============================================================================
# CONFIGURATION
# ============================================================================

def get_databridges_exchange_rate_path():
    """Get path to DataBridges Exchange Rate file"""
    if DATABRIDGES_BASE is None:
        return None
    return DATABRIDGES_BASE / "Exchange Rates" / "Exchange Rate - Data Bridges.xlsx"

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def export_databridges_exchange_rate(year: int, month: int):
    """
    Export exchange rate to DataBridges format (non-interactive, overwrites existing month)
    
    Appends or updates row with:
    - Year: Year number
    - Month: Month number
    - Date: Datetime object (will be formatted as M/D/YYYY in Excel)
    - Market: "Tripoli center"
    - official: Average official rate for the month
    - parallel: Average parallel rate for the month
    """
    
    print("="*70)
    print("LIBYA PMM - DATABRIDGES EXCHANGE RATE EXPORT")
    print("="*70)
    
    # Get paths from config
    paths = get_month_paths(year, month)
    month_name = paths["month_name"]
    
    print(f"\n📅 Processing: {month_name} {year}")
    
    # Input file (monthly exchange rate from OneDrive)
    input_file = paths["exchange_rate_monthly"]
    
    print(f"\n📂 Reading monthly exchange rate:")
    print(f"   {input_file}")
    
    # Check if file exists
    if not input_file.exists():
        print(f"\n❌ Error: Monthly exchange rate file not found!")
        print(f"   Expected location: {input_file}")
        print(f"\n   Make sure you have:")
        print(f"   1. Run process_exchange_rate.py")
        print(f"   2. Filled parallel market data in Excel")
        return False
    
    # Read monthly exchange rate
    monthly_df = pd.read_excel(input_file)
    print(f"   ✅ Loaded {len(monthly_df)} daily records")
    
    # Calculate averages
    print(f"\n📊 Calculating monthly averages...")
    
    # Official rate average
    official_avg = monthly_df["USD/LYD"].mean()
    print(f"   Official rate average: {official_avg:.4f}")
    
    # Parallel rate average (handle empty/missing values)
    parallel_col = monthly_df["Parallel Market USD/LYD"]
    parallel_col = parallel_col.replace("", float("nan"))
    parallel_numeric = pd.to_numeric(parallel_col, errors="coerce")
    
    if parallel_numeric.notna().any():
        parallel_avg = parallel_numeric.mean()
        print(f"   Parallel rate average: {parallel_avg:.4f}")
    else:
        print(f"   ⚠️  Warning: No parallel market data found")
        parallel_avg = float("nan")
    
    # Get DataBridges Exchange Rate file path
    databridges_file = get_databridges_exchange_rate_path()
    
    print(f"\n📂 DataBridges Exchange Rate file:")
    print(f"   {databridges_file}")
    
    # Read existing DataBridges file
    if databridges_file.exists():
        existing_df = pd.read_excel(databridges_file)
        print(f"   ✅ Loaded existing file ({len(existing_df)} records)")
        
        # Check if this month already exists
        existing_entry = existing_df[
            (existing_df["Year"] == year) &
            (existing_df["Month"] == month)
        ]
        
        if not existing_entry.empty:
            print(f"\n   ⚠️  Entry for {month_name} {year} already exists!")
            print("   Existing values:")
            print(f"      Official: {existing_entry['official'].iloc[0]}")
            print(f"      Parallel: {existing_entry['parallel'].iloc[0]}")
            print("   ➜ Overwriting existing entry with new values (non-interactive mode).")
            
            # Remove existing entry so we can insert new values cleanly
            existing_df = existing_df[
                ~((existing_df["Year"] == year) & (existing_df["Month"] == month))
            ]
            print("   Removed existing entry")
    else:
        print(f"   ⚠️  File doesn't exist, will create new file")
        existing_df = pd.DataFrame(
            columns=["Year", "Month", "Date", "Market", "official", "parallel"]
        )
    
    # Create new row with proper datetime object
    # Use first day of the month (will be formatted as M/D/YYYY in Excel)
    date_obj = datetime(year, month, 1)
    
    new_row = pd.DataFrame({
        "Year": [year],
        "Month": [month],
        "Date": [date_obj],  # Datetime object, not string
        "Market": ["Tripoli center"],
        "official": [round(official_avg, 2)],
        "parallel": [round(parallel_avg, 2) if not pd.isna(parallel_avg) else ""],
    })
    
    # Append new row
    updated_df = pd.concat([existing_df, new_row], ignore_index=True)
    
    # Sort by year and month
    updated_df = updated_df.sort_values(["Year", "Month"]).reset_index(drop=True)
    
    print(f"\n💾 Saving updated file...")
    
    # Ensure directory exists
    databridges_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Save updated file
    updated_df.to_excel(databridges_file, index=False, engine="openpyxl")
    
    # Apply date formatting to the Date column
    print(f"   🎨 Applying date formatting (displays as Month Year, value is M/D/YYYY)...")
    wb = load_workbook(databridges_file)
    ws = wb.active
    
    # Find the Date column (should be column C)
    date_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Date":
            date_col_idx = idx
            break
    
    if date_col_idx:
        # Apply "MMMM YYYY" format to all cells in Date column (skip header)
        # This displays as "November 2025" but stores as proper date
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=date_col_idx)
            if cell.value:
                # Set number format to display "Month Year" (e.g., "November 2025")
                # But the underlying value remains a proper date (11/1/2025)
                cell.number_format = 'MMMM YYYY'
    
    wb.save(databridges_file)
    
    print(f"   ✅ Saved {len(updated_df)} total records")
    print(f"   File size: {databridges_file.stat().st_size / 1024:.1f} KB")
    
    print("\n" + "="*70)
    print("✅ EXPORT COMPLETE")
    print("="*70)
    print(f"\n📤 Added {month_name} {year} to DataBridges Exchange Rate file")
    print(f"   Date display: {month_name} {year}")
    print(f"   Date value: {month}/1/{year} (proper date for sorting/filtering)")
    print(f"   Official: {official_avg:.2f}")
    if not pd.isna(parallel_avg):
        print(f"   Parallel: {parallel_avg:.2f}")
    else:
        print("   Parallel: N/A")
    
    return True

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Export exchange rate to DataBridges format (appends/updates existing file)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
    Examples:
    # Export November 2025 exchange rate
    python scripts/05_Data_Outputs/export_databridges_exchangerate.py 2025 11
    
    # Export October 2025 exchange rate
    python scripts/05_Data_Outputs/export_databridges_exchangerate.py 2025 10

    How it works:
    1. Reads monthly exchange rate file from OneDrive
    2. Calculates average official and parallel rates
    3. Appends or overwrites the row in the DataBridges Exchange Rate file
    4. Formats Date column to display as "Month Year" (e.g., "November 2025")
        - Display: "November 2025" 
        - Underlying value: 11/1/2025 (proper date for sorting)
    5. No interactive prompts (safe for automation)

    Prerequisites:
    - Monthly exchange rate file created and filled
    - File location: Monthly Reports/YYYY/MonthName/Exchange Rate/Exchange_Rate_MonthYY.xlsx
        """
    )
    
    parser.add_argument("year", type=int, help="Year (e.g., 2025)")
    parser.add_argument("month", type=int, help="Month (1-12)")
    
    args = parser.parse_args()
    
    # Validate month
    if not 1 <= args.month <= 12:
        print(f"❌ Error: Month must be between 1 and 12, got {args.month}")
        sys.exit(1)
    
    try:
        success = export_databridges_exchange_rate(args.year, args.month)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)