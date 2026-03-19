"""
Libya PMM - Generate Progress Tracker
Creates progress tracking Excel file with exact formatting specifications

Usage:
    python scripts/02_QAQC/generate_progress_tracker.py <year> <month>

Example:
    python scripts/02_QAQC/generate_progress_tracker.py 2025 12
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
import pytz
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/02_QAQC/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

from config import get_month_paths

# Commodity mappings
COMMODITY_LABELS = {
    'q_salt_price_per_kilo': 'Salt (Kg)',
    'q_sugar_price_per_kilo': 'Sugar (Kg)',
    'q_flour_price_per_kilo': 'Flour (Kg)',
    'q_rice_price_per_kilo': 'Rice (Kg)',
    'q_pasta_price_per_500g': 'Pasta (500g)',
    'q_couscous_price_per_kilo': 'Couscous (Kg)',
    'q_tomatop_price_per_400g': 'Tomato Paste (400g)',
    'q_chickpeas_price_per_400g': 'Chickpeas (400g)',
    'q_beans_price_per_400g': 'Beans (400g)',
    'q_cmilk_price_per_200ml': 'Condensed Milk (200ml)',
    'q_milk_price_per_liter': 'Milk (L)',
    'q_bmilk_price_per_400g': 'Baby Milk (400g)',
    'q_gtea_price_per_250g': 'Green Tea (250g)',
    'q_btea_price_per_250g': 'Black Tea (250g)',
    'q_oil_price_per_liter': 'Veg Oil (L)',
    'q_tuna_price_per_200g': 'Tuna (200g)',
    'q_eggs_price_per_30eggs': 'Eggs (30pc)',
    'q_chicken_price_per_kilo': 'Chicken (Kg)',
    'q_lamb_price_per_kilo': 'Lamb meat (Kg)',
    'q_bread_price_per_5medium_pieces': 'Bread (5pc)',
    'q_tomatoes_price_per_kilo': 'Tomato (Kg)',
    'q_onions_price_per_kilo': 'Onion (Kg)',
    'q_pepper_price_per_kilo': 'Peppers (Kg)',
    'q_potatoes_price_per_kilo': 'Potatoes (Kg)',
    'q_hwsoap_price_per_piece': 'Handwash Soap (Pc)',
    'q_lsoap_price_per_kilo': 'Laundry Powder (Kg)',
    'q_shampoo_price_per_250ml': 'Shampo (250ml)',
    'q_dsoap_price_per_liter': 'Dishwashing Liquid (L)',
    'q_ldet_price_per_litre': 'Laundry Detergent (L)',
    'q_toothpaste_price_per_tube': 'Toothpaste (Pc)',
    'q_toothbrush_price_per_piece': 'Toothbrush (Pc)',
    'q_spads_price_per_10pads': 'Sanitary Pads (10Pc)',
    'q_childrensdiapers_price_per_pack': 'Hand Sanitiser (L)',
    'q_water_price_per_15l': 'Sanitiser (L)',
    'q_fuel_public_price_per_11kg': 'Public fuel (11kg)',
    'q_public_gasoline_price_per_liter': 'Public Gasoline (L)',
    'q_private_gasoline_price_per_liter': 'Private Gasoline (L)',
    'q_fuel_private_price_per_11kg': 'Private fuel (11kg)',
}

FOOD_ITEMS = [
    'q_salt_price_per_kilo', 'q_sugar_price_per_kilo', 'q_flour_price_per_kilo',
    'q_rice_price_per_kilo', 'q_pasta_price_per_500g', 'q_couscous_price_per_kilo',
    'q_tomatop_price_per_400g', 'q_chickpeas_price_per_400g', 'q_beans_price_per_400g',
    'q_cmilk_price_per_200ml', 'q_milk_price_per_liter', 'q_bmilk_price_per_400g',
    'q_gtea_price_per_250g', 'q_btea_price_per_250g', 'q_oil_price_per_liter',
    'q_tuna_price_per_200g', 'q_eggs_price_per_30eggs', 'q_chicken_price_per_kilo',
    'q_lamb_price_per_kilo', 'q_bread_price_per_5medium_pieces', 'q_tomatoes_price_per_kilo',
    'q_onions_price_per_kilo', 'q_pepper_price_per_kilo', 'q_potatoes_price_per_kilo'
]

NON_FOOD_ITEMS = [
    'q_hwsoap_price_per_piece', 'q_lsoap_price_per_kilo', 'q_shampoo_price_per_250ml',
    'q_dsoap_price_per_liter', 'q_ldet_price_per_litre', 'q_toothpaste_price_per_tube',
    'q_toothbrush_price_per_piece', 'q_spads_price_per_10pads', 'q_childrensdiapers_price_per_pack',
    'q_water_price_per_15l'
]

FUEL_ITEMS = [
    'q_fuel_public_price_per_11kg', 'q_public_gasoline_price_per_liter',
    'q_private_gasoline_price_per_liter', 'q_fuel_private_price_per_11kg'
]


def count_prices_by_municipality(df: pd.DataFrame) -> pd.DataFrame:
    """Count non-null prices per commodity per municipality"""
    price_cols = [col for col in COMMODITY_LABELS.keys() if col in df.columns]
    if not price_cols:
        raise ValueError("No price columns found")
    if 'S1_06' not in df.columns:
        raise ValueError("Municipality column 'S1_06' not found")
    
    return df.groupby(['S1_06'])[price_cols].apply(lambda g: g.notnull().sum())


def create_excel_with_formatting(grouped_counts: pd.DataFrame, output_path: Path, 
                                 month_name: str, year: int):
    """Create formatted Excel matching exact specifications"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, GradientFill, Border, Side
    except ImportError:
        print("⚠️  openpyxl not available")
        return
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get Libya timezone timestamp
    libya_tz = pytz.timezone('Africa/Tripoli')
    libya_time = datetime.now(libya_tz)
    timestamp = libya_time.strftime("%Y-%m-%d @ %H:%M:%S GMT+2")
    
    # Define borders
    light_gray_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    thick_top_bottom = Border(
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    for category_name, category_cols in [
        ('Food Items', FOOD_ITEMS),
        ('Non-Food Items', NON_FOOD_ITEMS),
        ('Fuel Items', FUEL_ITEMS)
    ]:
        existing_cols = [col for col in category_cols if col in grouped_counts.columns]
        if not existing_cols:
            continue
        
        ws = wb.create_sheet(category_name)
        
        # Row 2: Title (size 16, bold, black)
        ws['A2'] = f'PMM - {month_name} {year}'
        ws['A2'].font = Font(name='Aptos Narrow', size=16, bold=True, color='000000')
        ws.merge_cells('A2:Z2')
        
        # Row 3: Subtitle (size 11, bold, gray #747474)
        ws['A3'] = f'{category_name} - Data Collection Status'
        ws['A3'].font = Font(name='Aptos Narrow', size=11, bold=True, color='747474')
        ws.merge_cells('A3:Z3')
        
        # Row 4: Timestamp (size 9, bold italic, blue #0070C0)
        ws['A4'] = f'* Last update: {timestamp}'
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, italic=True, color='0070C0')
        ws.merge_cells('A4:Z4')
        
        # Row 6: "Summary" (size 16, bold, black)
        ws['A6'] = 'Summary'
        ws['A6'].font = Font(name='Aptos Narrow', size=16, bold=True, color='000000')
        
        # Calculate metrics
        cat_data = grouped_counts[existing_cols]
        
        # Unique Price Points: Count only up to 4 per cell
        unique_prices = int((cat_data.clip(upper=4)).sum().sum())
        
        total_required = len(cat_data) * len(existing_cols) * 4
        completion_pct = (unique_prices / total_required * 100) if total_required > 0 else 0
        
        # KPI Card 1: Unique Market Price Points (3 columns wide: A-C)
        ws.merge_cells('A7:C13')
        ws['A7'] = unique_prices
        ws['A7'].font = Font(name='Aptos Narrow', size=36, bold=False, color='000000')
        ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A7'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        ws.merge_cells('A14:C15')
        ws['A14'] = 'Unique Market Price Points'
        ws['A14'].font = Font(name='Aptos Narrow', size=11, bold=False, color='000000')
        ws['A14'].alignment = Alignment(horizontal='center', vertical='center')
        
        # KPI Card 2: Total Price Points Required (4 columns wide: D-G)
        ws.merge_cells('D7:G13')
        ws['D7'] = total_required
        ws['D7'].font = Font(name='Aptos Narrow', size=36, bold=False, color='000000')
        ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D7'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        ws.merge_cells('D14:G15')
        ws['D14'] = 'Total Price Points Required'
        ws['D14'].font = Font(name='Aptos Narrow', size=11, bold=False, color='000000')
        ws['D14'].alignment = Alignment(horizontal='center', vertical='center')
        
        # KPI Card 3: Completion % (4 columns wide: H-K)
        ws.merge_cells('H7:K13')
        ws['H7'] = f'{completion_pct:.1f}%'
        ws['H7'].font = Font(name='Aptos Narrow', size=36, bold=False, color='000000')
        ws['H7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H7'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        ws.merge_cells('H14:K15')
        ws['H14'] = 'Completion %'
        ws['H14'].font = Font(name='Aptos Narrow', size=11, bold=False, color='000000')
        ws['H14'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 16: Empty
        
        # Row 17: Table headers (background #2D5E7F, white text, size 11, bold)
        cat_data_display = cat_data.reset_index()
        cat_data_display = cat_data_display.rename(columns={'S1_06': ''})
        cat_data_display = cat_data_display.rename(columns=COMMODITY_LABELS)
        
        # Set header row height to make it taller (default is ~15, increase to 30)
        ws.row_dimensions[17].height = 30
        
        for col_idx, col_name in enumerate(cat_data_display.columns, 1):
            cell = ws.cell(row=17, column=col_idx, value=col_name)
            cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2D5E7F', end_color='2D5E7F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 18+: Data with gradient color coding
        for row_idx, row_data in enumerate(cat_data_display.itertuples(index=False), 18):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Aptos Narrow', size=11, color='000000')
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # Apply gradient fill based on value
                if col_idx > 1 and isinstance(value, (int, float)):
                    if value > 4:
                        # Dark green
                        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                    elif value == 4:
                        # Light green
                        cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                    elif value == 3:
                        # Light pink
                        cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                    elif value == 2:
                        # Pink
                        cell.fill = PatternFill(start_color='E7959B', end_color='E7959B', fill_type='solid')
                    elif value == 1:
                        # Light red
                        cell.fill = PatternFill(start_color='D07378', end_color='D07378', fill_type='solid')
                    elif value == 0:
                        # Dark red
                        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                        cell.font = Font(name='Aptos Narrow', size=11, color='FFFFFF')
        
        # Grand Total row - immediately after last data row (no empty row)
        last_data_row = 18 + len(cat_data_display) - 1
        total_row = last_data_row + 1  # Changed from +2 to +1 (no empty row)
        
        # Apply thick borders to all cells in grand total row
        for col_idx in range(1, len(cat_data_display.columns) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.border = thick_top_bottom
        
        ws.cell(row=total_row, column=1, value='Grand Total')
        ws.cell(row=total_row, column=1).font = Font(name='Aptos Narrow', size=11, bold=True, color='000000')
        ws.cell(row=total_row, column=1).fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
        
        for col_idx in range(2, len(cat_data_display.columns) + 1):
            col_sum = int(cat_data_display.iloc[:, col_idx-1].sum())
            cell = ws.cell(row=total_row, column=col_idx, value=col_sum)
            cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='000000')
            cell.alignment = Alignment(horizontal='center')
            
            max_possible = len(cat_data_display) * 4
            if col_sum >= max_possible:
                # Use dark green #70AD47 for complete
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            elif col_sum < max_possible * 0.75:
                cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            else:
                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(cat_data_display.columns) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 12
    
    # --- Missing Prices Sheet ---
    ws_missing = wb.create_sheet('Missing Prices')
    
    # Commodities to exclude from missing prices
    EXCLUDE_FROM_MISSING = [
        'q_private_gasoline_price_per_liter',
        'q_fuel_private_price_per_11kg'
    ]
    
    # Create missing prices by category
    current_row = 1
    
    for category_name, category_cols in [
        ('Food Items', FOOD_ITEMS),
        ('Non-Food Items', NON_FOOD_ITEMS),
        ('Fuel Items', FUEL_ITEMS)
    ]:
        # Filter: only include columns that exist AND are not excluded
        filtered_cols = [col for col in category_cols 
                        if col in grouped_counts.columns 
                        and col not in EXCLUDE_FROM_MISSING]
        
        if not filtered_cols:
            continue
        
        # Collect missing data for this category ONLY
        category_missing = []
        for municipality in grouped_counts.index:
            for col in filtered_cols:
                collected = grouped_counts.loc[municipality, col]
                if collected < 4:
                    commodity_label = COMMODITY_LABELS.get(col, col)
                    missing = 4 - collected
                    category_missing.append({
                        'Municipality': municipality,
                        'Commodity': commodity_label,
                        'Collected': int(collected),
                        'Missing': int(missing)
                    })
        
        if not category_missing:
            continue
        
        # Category title
        title_cell = ws_missing.cell(row=current_row, column=1, value=category_name)
        title_cell.font = Font(name='Aptos Narrow', size=14, bold=True, color='000000')
        current_row += 1
        
        # Headers
        headers = ['Municipality', 'Commodity', 'Collected', 'Missing']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_missing.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2D5E7F', end_color='2D5E7F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1
        
        # Sort and write data
        missing_df = pd.DataFrame(category_missing)
        missing_df = missing_df.sort_values(['Municipality', 'Commodity']).reset_index(drop=True)
        
        for row_data in missing_df.itertuples(index=False):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_missing.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(name='Aptos Narrow', size=11, color='000000')
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Empty row between categories
        current_row += 2
    
    # Auto-size columns
    ws_missing.column_dimensions['A'].width = 20
    ws_missing.column_dimensions['B'].width = 30
    ws_missing.column_dimensions['C'].width = 12
    ws_missing.column_dimensions['D'].width = 12
    
    wb.save(output_path)


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate progress tracker")
    parser.add_argument('year', type=int)
    parser.add_argument('month', type=int)
    args = parser.parse_args()
    
    if not 1 <= args.month <= 12:
        print("❌ Error: Month must be 1-12")
        sys.exit(1)
    
    paths = get_month_paths(args.year, args.month)
    month_name = paths['month_name']
    month_tag = paths['month_tag']
    
    print("="*70)
    print("GENERATE PROGRESS TRACKER")
    print("="*70)
    print(f"\nTarget: {month_name} {args.year}")
    
    input_file = paths['raw_data']
    if not input_file.exists():
        print(f"\n❌ File not found: {input_file}")
        sys.exit(1)
    
    print(f"\n📂 Reading: {input_file.name}")
    df = pd.read_excel(input_file)
    print(f"   ✓ {len(df)} records")
    
    print(f"\n🔢 Counting prices...")
    grouped_counts = count_prices_by_municipality(df)
    print(f"   ✓ {len(grouped_counts)} municipalities, {len(grouped_counts.columns)} commodities")
    
    output_file = paths['progress'] / f"PMM Progress Tracker - {month_name} {str(args.year)[-2:]}.xlsx"
    paths['progress'].mkdir(parents=True, exist_ok=True)
    
    print(f"\n💾 Creating: {output_file.name}")
    create_excel_with_formatting(grouped_counts, output_file, month_name, args.year)
    print(f"   ✓ Done")
    
    print(f"\n✅ File: {output_file}")
    print()


if __name__ == "__main__":
    main()