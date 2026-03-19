"""
Libya PMM - Generate Follow-Up File
Creates follow-up file with outlier detection and boxplot visualizations

This script:
1. Detects price outliers using IQR method
2. Creates Follow-Up table with prices to be verified
3. Generates boxplot visualizations for all commodities

Usage:
    python scripts/02_QAQC/generate_followup.py <year> <month>

Example:
    python scripts/02_QAQC/generate_followup.py 2025 12
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import pytz
import os

# Auto-detect environment and set project root
if os.path.exists('/app'):  # Running in Docker
    PROJECT_ROOT = Path('/app')
else:  # Running locally
    # Script is in scripts/02_QAQC/
    PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Add project root to path to import config
sys.path.insert(0, str(PROJECT_ROOT))

from config import get_month_paths

# Commodity labels (same as progress tracker)
COMMODITY_LABELS = {
    'q_salt_price_per_kilo': 'Salt (Kg)',
    'q_sugar_price_per_kilo': 'Sugar (Kg)',
    'q_flour_price_per_kilo': 'Flour (Kg)',
    'q_rice_price_per_kilo': 'Rice (Kg)',
    'q_pasta_price_per_500g': 'Pasta (500g)',
    'q_couscous_price_per_kilo': 'Couscous (Kg)',
    'q_tomatop_price_per_400g': 'Tomato Paste (400g)',
    'q_chickpeas_price_per_400g': 'Chickpeas (400g)',
    'q_beans_price_per_400g': 'Beans (400g)',
    'q_cmilk_price_per_200ml': 'Condensed Milk (200ml)',
    'q_milk_price_per_liter': 'Milk (L)',
    'q_bmilk_price_per_400g': 'Baby Milk (400g)',
    'q_gtea_price_per_250g': 'Green Tea (250g)',
    'q_btea_price_per_250g': 'Black Tea (250g)',
    'q_oil_price_per_liter': 'Veg Oil (L)',
    'q_tuna_price_per_200g': 'Tuna (200g)',
    'q_eggs_price_per_30eggs': 'Eggs (30pc)',
    'q_chicken_price_per_kilo': 'Chicken (Kg)',
    'q_lamb_price_per_kilo': 'Lamb meat (Kg)',
    'q_bread_price_per_5medium_pieces': 'Bread (5pc)',
    'q_tomatoes_price_per_kilo': 'Tomato (Kg)',
    'q_onions_price_per_kilo': 'Onion (Kg)',
    'q_pepper_price_per_kilo': 'Peppers (Kg)',
    'q_potatoes_price_per_kilo': 'Potatoes (Kg)',
    'q_hwsoap_price_per_piece': 'Handwash Soap (Pc)',
    'q_lsoap_price_per_kilo': 'Laundry Powder (Kg)',
    'q_shampoo_price_per_250ml': 'Shampo (250ml)',
    'q_dsoap_price_per_liter': 'Dishwashing Liquid (L)',
    'q_ldet_price_per_litre': 'Laundry Detergent (L)',
    'q_toothpaste_price_per_tube': 'Toothpaste (Pc)',
    'q_toothbrush_price_per_piece': 'Toothbrush (Pc)',
    'q_spads_price_per_10pads': 'Sanitary Pads (10Pc)',
    'q_childrensdiapers_price_per_pack': 'Hand Sanitiser (L)',
    'q_water_price_per_15l': 'Sanitiser (L)',
    'q_fuel_public_price_per_11kg': 'Public fuel (11kg)',
    'q_public_gasoline_price_per_liter': 'Public Gasoline (L)',
    'q_private_gasoline_price_per_liter': 'Private Gasoline (L)',
    'q_fuel_private_price_per_11kg': 'Private fuel (11kg)',
}


def detect_outliers_iqr(series):
    """Identify outliers using IQR method"""
    q1 = series.quantile(0.25)
    q3 = series.quantile(0.75)
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    return (series < lower) | (series > upper)


def detect_all_outliers(df):
    """Detect outliers in all price columns"""
    # Get all columns with '_price_per_'
    price_per_cols = [col for col in df.columns if '_price_per_' in col]
    
    outlier_records = []
    
    # Check for required columns
    if '_id' not in df.columns:
        raise ValueError("_id column not found in data")
    
    if '_xform_id' not in df.columns:
        raise ValueError("_xform_id column not found in data")
    
    for outlier_col in price_per_cols:
        # Base name like 'q_salt'
        base = outlier_col.split('_price_per_')[0]
        
        # Find corresponding columns
        original_col = base + '_price'
        quantity1_col = base + '_quantity1'
        quantity2_col = base + '_quantity2'
        
        if original_col not in df.columns:
            continue
        
        # Detect outliers
        outliers = detect_outliers_iqr(df[outlier_col])
        
        if not outliers.any():
            continue
        
        # Extract relevant data for outliers - use _id specifically
        cols_to_extract = [
            'S1_01',        # Enumerator name
            'S1_06',        # Municipality
            'q_shop_type',  # Shop Type
            '_id',          # Instance ID - use this specifically
            '_xform_id',    # Form ID from data
            original_col,
            outlier_col
        ]
        
        # Extract relevant data for outliers
        temp_df = df[outliers][cols_to_extract].copy()
        
        # Add commodity and quantity info
        temp_df['Commodity'] = COMMODITY_LABELS.get(outlier_col, outlier_col)
        temp_df['original_value'] = df.loc[outliers, original_col]
        temp_df['outlier_column'] = outlier_col
        temp_df['outlier_value'] = df.loc[outliers, outlier_col]
        temp_df['quantity1'] = df.loc[outliers, quantity1_col] if quantity1_col in df.columns else np.nan
        temp_df['quantity2'] = df.loc[outliers, quantity2_col] if quantity2_col in df.columns else np.nan
        
        # Keep only final columns
        temp_df = temp_df[[
            'S1_01', 'S1_06', 'q_shop_type',
            'Commodity', 'original_value',
            'outlier_column', 'outlier_value',
            'quantity1', 'quantity2', '_id', '_xform_id'
        ]]
        
        outlier_records.append(temp_df)
    
    if outlier_records:
        return pd.concat(outlier_records, ignore_index=True)
    else:
        return pd.DataFrame()


def create_followup_table(df_outliers):
    """Create formatted follow-up table"""
    if df_outliers.empty:
        return pd.DataFrame()
    
    # Create Form Link column using _xform_id and _id from data
    base_url = "https://moda.wfp.org/lby/26549/"
    df_outliers['Form Link'] = df_outliers.apply(
        lambda row: f"{base_url}{row['_xform_id']}/webform?instance-id={row['_id']}", 
        axis=1
    )
    
    # Rename columns
    rename_dict = {
        'S1_01': 'Enumerator Name',
        'S1_06': 'Municipality',
        'q_shop_type': 'Shop Type',
        'Commodity': 'Commodity',
        'original_value': 'Price (To Be Verified)',
        'outlier_column': 'Column',
        'outlier_value': 'Outlier Price',
        'quantity1': 'Available in Original Quantity?',
        'quantity2': 'If no, other Quantity?',
        '_id': 'Form Instance ID',
        'Form Link': 'Form Link'
    }
    
    df_readable = df_outliers.rename(columns=rename_dict).copy()
    
    # Drop outlier price and _xform_id columns
    df_readable.drop(columns=['Outlier Price', '_xform_id'], inplace=True, errors='ignore')
    
    # Add empty column for corrections
    df_readable['Correct Value'] = ""
    
    # Add index column (starting from 1)
    df_readable.insert(0, '#', range(1, len(df_readable) + 1))
    
    # Define column order
    ordered_cols = [
        '#',
        'Form Instance ID',
        'Form Link',
        'Enumerator Name',
        'Municipality',
        'Shop Type',
        'Commodity',
        'Column',
        'Available in Original Quantity?',
        'If no, other Quantity?',
        'Price (To Be Verified)',
        'Correct Value'
    ]
    
    # Reorder columns
    remaining_cols = [col for col in df_readable.columns if col not in ordered_cols]
    df_readable = df_readable[ordered_cols + remaining_cols]
    
    return df_readable


def create_excel_with_formatting(followup_table, output_path, month_name, year):
    """Create formatted Excel file with Follow-Up table"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  openpyxl not available")
        # Fallback to basic Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            followup_table.to_excel(writer, sheet_name='Follow Up', index=False)
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Follow Up'
    
    # Get Libya timezone timestamp
    libya_tz = pytz.timezone('Africa/Tripoli')
    libya_time = datetime.now(libya_tz)
    timestamp = libya_time.strftime("%Y-%m-%d @ %H:%M:%S GMT+2")
    
    # Row 2: Title (size 16, bold, black)
    ws['A2'] = f'PMM - {month_name} {year}'
    ws['A2'].font = Font(name='Aptos Narrow', size=16, bold=True, color='000000')
    ws.merge_cells('A2:K2')
    
    # Row 3: Subtitle (size 11, bold, gray #747474)
    ws['A3'] = 'QAQC - Data Validation Follow-Up'
    ws['A3'].font = Font(name='Aptos Narrow', size=11, bold=True, color='747474')
    ws.merge_cells('A3:K3')
    
    # Row 4: Timestamp (size 9, bold italic, blue #0070C0)
    ws['A4'] = f'* Last update: {timestamp}'
    ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, italic=True, color='0070C0')
    ws.merge_cells('A4:K4')
    
    # Row 5: Empty
    
    # Row 6: Headers
    headers = list(followup_table.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Special colors for specific columns
        if header == 'Price (To Be Verified)':
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        elif header == 'Correct Value':
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        else:
            # Default header color (dark blue)
            cell.fill = PatternFill(start_color='2D5E7F', end_color='2D5E7F', fill_type='solid')
    
    # Row 7+: Data
    for row_idx, row_data in enumerate(followup_table.itertuples(index=False), 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Aptos Narrow', size=11, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Get header for this column
            header = headers[col_idx - 1]
            
            # Make Form Link clickable
            if header == 'Form Link' and value:
                cell.hyperlink = value
                cell.font = Font(name='Aptos Narrow', size=11, color='0563C1', underline='single')
            
            # Color code columns with very light colors
            if header == 'Price (To Be Verified)':
                # Very light red
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif header == 'Correct Value':
                # Very light green
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Auto-size columns based on content
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        
        # Calculate max width needed
        max_length = len(str(header))
        
        # Special handling for Form Link - set to show full URL
        if header == 'Form Link':
            ws.column_dimensions[col_letter].width = 60  # Wide enough for full link
            continue
        
        # Special handling for index column
        if header == '#':
            ws.column_dimensions[col_letter].width = 5
            continue
        
        for row in ws.iter_rows(min_row=7, max_row=6 + len(followup_table), 
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        # Set column width (add padding)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Follow-Up file with outlier detection")
    parser.add_argument('year', type=int, help='Year (e.g., 2025)')
    parser.add_argument('month', type=int, help='Month (1-12)')
    
    args = parser.parse_args()
    
    if not 1 <= args.month <= 12:
        print("❌ Error: Month must be 1-12")
        sys.exit(1)
    
    paths = get_month_paths(args.year, args.month)
    month_name = paths['month_name']
    month_tag = paths['month_tag']
    
    print("="*70)
    print("GENERATE FOLLOW-UP FILE")
    print("="*70)
    print(f"\nTarget: {month_name} {args.year}")
    
    # Input file
    input_file = paths['raw_data']
    if not input_file.exists():
        print(f"\n❌ File not found: {input_file}")
        sys.exit(1)
    
    print(f"\n📂 Reading: {input_file.name}")
    df = pd.read_excel(input_file)
    print(f"   ✓ {len(df)} records")
    
    # Detect outliers
    print(f"\n🔍 Detecting outliers...")
    df_outliers = detect_all_outliers(df)
    
    if df_outliers.empty:
        print("   ✓ No outliers detected!")
        print("\n🎉 All prices look good - no follow-up needed")
        return
    
    print(f"   ⚠️  Found {len(df_outliers)} outlier records")
    
    # Create follow-up table
    print(f"\n📋 Creating follow-up table...")
    followup_table = create_followup_table(df_outliers)
    print(f"   ✓ {len(followup_table)} items to verify")
    
    # Output file - save to Progress folder
    excel_file = paths['progress'] / f"PMM QAQC Follow-Up - {month_name} {args.year}.xlsx"
    
    print(f"\n💾 Creating Excel file...")
    create_excel_with_formatting(followup_table, excel_file, month_name, args.year)
    print(f"   ✓ Follow-Up file saved")
    
    print("\n" + "="*70)
    print("✅ FOLLOW-UP FILE CREATED")
    print("="*70)
    print(f"\nFile created:")
    print(f"  • {excel_file}")
    print(f"    - Follow Up sheet: {len(followup_table)} items to verify")
    print(f"\n⚠️  {len(followup_table)} prices require verification")
    print()


if __name__ == "__main__":
    main()